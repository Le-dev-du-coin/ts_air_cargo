from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Count, Sum, Q, Avg, Case, When, IntegerField
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from django.db import transaction
from django.template.loader import render_to_string
from django.conf import settings
import os
import json
import datetime
from xhtml2pdf import pisa
from io import BytesIO
import pandas as pd
from django.db.models import F, Value, CharField
from django.db.models.functions import Concat
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from .models import Depense, ReceptionLot, Livraison, PriceAdjustment
from agent_chine_app.models import Lot, Colis, Client
from notifications_app.services import NotificationService
from django.contrib.auth import get_user_model

User = get_user_model()

# Décorateur pour vérifier que l'utilisateur est un agent mali
def agent_mali_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or not request.user.is_agent_mali:
            messages.error(request, "Accès refusé. Vous devez être un agent au Mali.")
            return redirect('authentication:login_agent_mali')
        return view_func(request, *args, **kwargs)
    return wrapper

@agent_mali_required
def exporter_lot_pdf(request, lot_id):
    """
    Exporte les détails d'un lot en PDF
    """
    lot = get_object_or_404(Lot, id=lot_id)
    colis_list = lot.colis.all().order_by('date_creation')
    
    # Statistiques
    total_colis = colis_list.count()
    colis_livres = colis_list.filter(statut='livre').count()
    colis_perdus = colis_list.filter(statut='perdu').count()
    colis_en_attente = total_colis - colis_livres - colis_perdus
    
    context = {
        'lot': lot,
        'colis_list': colis_list,
        'total_colis': total_colis,
        'colis_livres': colis_livres,
        'colis_perdus': colis_perdus,
        'colis_en_attente': colis_en_attente,
        'date_export': timezone.now().strftime('%d/%m/%Y %H:%M')
    }
    
    # Rendu du template en HTML
    html_string = render_to_string('agent_mali_app/export_lot_pdf.html', context)
    
    # Création du PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="lot_{lot.numero_lot}_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf"'
    
    # Génération du PDF
    pdf = pisa.CreatePDF(
        html_string,
        dest=response,
        encoding='UTF-8',
        link_callback=None
    )
    
    if not pdf.err:
        return response
    
    return HttpResponse('Une erreur est survenue lors de la génération du PDF', status=500)

def details_lot_view(request, lot_id):
    """
    Affiche les détails d'un lot spécifique
    """
    lot = get_object_or_404(Lot, id=lot_id)
    
    # Récupérer tous les colis du lot avec pagination
    colis_list = lot.colis.all().order_by('date_creation')
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(colis_list, 20)  # 20 colis par page
    
    try:
        colis = paginator.page(page)
    except PageNotAnInteger:
        colis = paginator.page(1)
    except EmptyPage:
        colis = paginator.page(paginator.num_pages)
    
    # Statistiques
    total_colis = colis_list.count()
    colis_livres = colis_list.filter(statut='livre').count()
    colis_perdus = colis_list.filter(statut='perdu').count()
    colis_en_attente = total_colis - colis_livres - colis_perdus
    
    context = {
        'lot': lot,
        'colis': colis,
        'total_colis': total_colis,
        'colis_livres': colis_livres,
        'colis_perdus': colis_perdus,
        'colis_en_attente': colis_en_attente,
        'title': f'Détails du lot {lot.numero_lot}'
    }
    
    return render(request, 'agent_mali_app/details_lot.html', context)

@agent_mali_required
def lots_livres_view(request):
    """
    Affiche la liste des lots complètement traités (tous les colis sont soit livrés, soit perdus)
    """
    # Récupérer la requête de recherche si elle existe
    query = request.GET.get('q', '')
    
    # Récupérer tous les lots
    lots = Lot.objects.all()
    
    # Filtrer les lots où tous les colis sont soit livrés, soit perdus
    lots_complets = []
    for lot in lots:
        total_colis = lot.colis.count()
        if total_colis == 0:
            continue  # Ignorer les lots sans colis
            
        colis_livres = lot.colis.filter(statut='livre').count()
        colis_perdus = lot.colis.filter(statut='perdu').count()
        
        # Vérifier si tous les colis sont soit livrés, soit perdus
        if (colis_livres + colis_perdus) == total_colis:
            lots_complets.append(lot)
    
    # Convertir en queryset pour la pagination
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    from django.db.models import Q
    
    # Filtrer par requête de recherche si nécessaire
    if query:
        lots_complets = [
            lot for lot in lots_complets 
            if (query.lower() in lot.numero_lot.lower()) or
               any(query.lower() in (colis.client.user.get_full_name() or '').lower() for colis in lot.colis.all())
        ]
    
    # Trier par date d'expédition (du plus récent au plus ancien)
    lots_complets.sort(key=lambda x: x.date_expedition or x.date_creation, reverse=True)
    
    # Calculer les statistiques
    total_lots = len(lots_complets)
    total_colis = sum(lot.colis.count() for lot in lots_complets)
    total_colis_livres = sum(lot.colis.filter(statut='livre').count() for lot in lots_complets)
    total_colis_perdus = sum(lot.colis.filter(statut='perdu').count() for lot in lots_complets)
    
    # Pagination
    paginator = Paginator(lots_complets, 20)  # 20 lots par page
    page = request.GET.get('page')
    
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    context = {
        'page_obj': page_obj,
        'query': query,
        'total_lots': total_lots,
        'total_colis': total_colis,
        'total_colis_livres': total_colis_livres,
        'total_colis_perdus': total_colis_perdus,
        'title': 'Lots Complètement Traités',
    }
    
    return render(request, 'agent_mali_app/liste_lots_livres.html', context)


def dashboard_view(request):
    """
    Tableau de bord pour Agent Mali avec statistiques temps réel
    """
    # Statistiques des lots - inclure les lots expédiés comme en transit pour l'agent Mali
    lots_en_transit = Lot.objects.filter(statut__in=['expedie', 'en_transit']).count()
    lots_arrives = Lot.objects.filter(statut='arrive').count()
    lots_expedies_total = Lot.objects.filter(statut='expedie').count()
    
    # Statistiques des colis - inclure les colis expédiés comme en transit pour l'agent Mali
    colis_en_transit = Colis.objects.filter(statut__in=['expedie', 'en_transit']).count()
    colis_arrives = Colis.objects.filter(statut='arrive').count()
    colis_livres = Colis.objects.filter(statut='livre').count()
    colis_perdus = Colis.objects.filter(statut='perdu').count()
    
    # Statistiques de livraison aujourd'hui
    aujourd_hui = timezone.now().date()
    livraisons_aujourd_hui = Livraison.objects.filter(
        date_livraison_effective__date=aujourd_hui,
        statut='livree'
    ).count()
    
    # Calculs de revenus et dépenses du mois
    debut_mois = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # Revenus basés sur le prix de transport des colis livrés ce mois
    revenus_livraison_mois = Colis.objects.filter(
        statut='livre',
        livraisons__date_livraison_effective__gte=debut_mois,
        livraisons__statut='livree'
    ).aggregate(total=Sum('prix_calcule'))['total'] or 0
    
    # Revenus journaliers (colis livrés aujourd'hui)
    revenus_livraison_aujourd_hui = Colis.objects.filter(
        statut='livre',
        livraisons__date_livraison_effective__date=aujourd_hui,
        livraisons__statut='livree'
    ).aggregate(total=Sum('prix_calcule'))['total'] or 0
    
    # Valeur totale des colis en stock (arrivés mais pas encore livrés)
    valeur_stock_magasin = Colis.objects.filter(
        statut='arrive'
    ).aggregate(total=Sum('prix_calcule'))['total'] or 0
    
    # Dépenses du mois
    depenses_mois = Depense.objects.filter(
        date_depense__gte=debut_mois.date()
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    # Calcul du bénéfice mensuel
    benefice_mois = float(revenus_livraison_mois) - float(depenses_mois)
    
    # Derniers lots reçus - inclure les lots expédiés et en transit
    derniers_lots_recus = Lot.objects.filter(
        statut__in=['arrive', 'en_transit', 'expedie']
    ).order_by('-date_expedition')[:5]
    
    # Dernières livraisons
    dernieres_livraisons = Livraison.objects.filter(
        statut='livree'
    ).select_related('colis__client__user').order_by('-date_livraison_effective')[:5]
    
    # Colis à livrer (priorité)
    colis_a_livrer = Colis.objects.filter(
        statut='arrive'
    ).select_related('client__user', 'lot').count()
    
    # Colis livrés en attente de paiement
    colis_attente_paiement = Colis.objects.filter(
        statut='livre',
        livraisons__statut='livree',
        livraisons__statut_paiement='en_attente'
    ).distinct().count()
    
    # Statistiques des ajustements de prix du mois
    ajustements_mois = PriceAdjustment.objects.filter(
        created_at__gte=debut_mois,
        status='applied'
    )
    
    # Montant des jetons cédés ce mois
    jetons_cedes_mois = ajustements_mois.filter(
        adjustment_type='jc'
    ).aggregate(total=Sum('adjustment_amount'))['total'] or 0
    
    # Montant des remises ce mois
    remises_mois = ajustements_mois.filter(
        adjustment_type='remise'
    ).aggregate(total=Sum('adjustment_amount'))['total'] or 0
    
    # Derniers ajustements
    derniers_ajustements = PriceAdjustment.objects.filter(
        applied_by__is_agent_mali=True
    ).select_related('colis__client__user', 'applied_by').order_by('-created_at')[:5]
    
    context = {
        'stats': {
            'lots_en_transit': lots_en_transit,
            'lots_arrives': lots_arrives,
            'lots_expedies_total': lots_expedies_total,
            'colis_en_transit': colis_en_transit,
            'colis_arrives': colis_arrives,
            'colis_livres': colis_livres,
            'colis_perdus': colis_perdus,
            'livraisons_aujourd_hui': livraisons_aujourd_hui,
            'revenus_livraison_mois': float(revenus_livraison_mois),
            'revenus_livraison_aujourd_hui': float(revenus_livraison_aujourd_hui),
            'valeur_stock_magasin': float(valeur_stock_magasin),
            'depenses_mois': float(depenses_mois),
            'benefice_mois': benefice_mois,
            'colis_a_livrer': colis_a_livrer,
            'colis_attente_paiement': colis_attente_paiement,
            # Statistiques ajustements
            'jetons_cedes_mois': float(jetons_cedes_mois),
            'remises_mois': float(remises_mois),
            'ajustements_total_mois': float(jetons_cedes_mois) + float(remises_mois),
        },
        'derniers_lots_recus': derniers_lots_recus,
        'dernieres_livraisons': dernieres_livraisons,
        'derniers_ajustements': derniers_ajustements,
    }
    return render(request, 'agent_mali_app/dashboard.html', context)

@agent_mali_required
def colis_attente_paiement_view(request):
    """
    Vue pour gérer les colis livrés en attente de paiement
    """
    colis_attente = Colis.objects.filter(
        statut='livre',
        livraisons__statut='livree',
        livraisons__statut_paiement='en_attente'
    ).select_related(
        'client__user', 'lot'
    ).prefetch_related('livraisons').distinct().order_by('-livraisons__date_livraison_effective')
    
    context = {
        'colis_attente': colis_attente,
        'title': 'Colis en Attente de Paiement',
    }
    return render(request, 'agent_mali_app/colis_attente_paiement.html', context)

@agent_mali_required
def gestion_paiement_lot_view(request, lot_id):
    """
    Vue pour gérer le paiement d'un lot complet
    """
    lot = get_object_or_404(Lot, id=lot_id)
    
    # Vérifier que l'utilisateur a les droits pour ce lot
    if not request.user.is_superuser and lot.agent_mali != request.user:
        messages.error(request, "Vous n'avez pas la permission de gérer ce lot.")
        return redirect('agent_mali:dashboard')
    
    # Récupérer les colis du lot avec leurs ajustements
    colis_list = lot.colis.all().prefetch_related('price_adjustments')
    
    # Calculer les totaux
    total_colis = colis_list.count()
    colis_livres = colis_list.filter(statut='livre').count()
    colis_avec_jc = colis_list.filter(price_adjustments__adjustment_type='jc', 
                                    price_adjustments__status='active').distinct().count()
    
    # Vérifier si tous les colis sont livrés
    if colis_livres < total_colis:
        messages.warning(
            request, 
            f"⚠️ Ce lot contient {total_colis - colis_livres} colis non encore livrés. "
            "Tous les colis doivent être livrés avant de procéder au paiement."
        )
        return redirect('agent_mali:details_lot', lot_id=lot.id)
    
    # Traitement du formulaire de paiement
    if request.method == 'POST':
        with transaction.atomic():
            # Marquer tous les colis comme payés
            for colis in colis_list:
                # Appliquer les ajustements de type Jeton Cédé s'ils existent
                jc_adjustments = colis.price_adjustments.filter(
                    adjustment_type='jc',
                    status='active'
                )
                
                # Appliquer chaque ajustement Jeton Cédé
                for adjustment in jc_adjustments:
                    try:
                        adjustment.apply_adjustment()
                        messages.info(
                            request,
                            f"ℹ️ Ajustement Jeton Cédé de {adjustment.adjustment_amount} FCFA "
                            f"appliqué au colis {colis.numero_suivi}."
                        )
                    except Exception as e:
                        logger.error(f"Erreur lors de l'application de l'ajustement {adjustment.id}: {str(e)}")
                
                # Mettre à jour le statut de paiement des livraisons
                livraisons = colis.livraisons.filter(statut='livree', statut_paiement='en_attente')
                livraisons.update(statut_paiement='paye')
                
                # Mettre à jour le statut du colis
                colis.statut = 'paye'
                colis.save()
            
            # Mettre à jour le statut du lot
            lot.statut = 'paye'
            lot.save()
            
            messages.success(
                request,
                f"✅ Paiement du lot {lot.numero_lot} enregistré avec succès. "
                f"{total_colis} colis marqués comme payés."
            )
            
            return redirect('agent_mali:details_lot', lot_id=lot.id)
    
    # Préparer le contexte pour le template
    context = {
        'lot': lot,
        'total_colis': total_colis,
        'colis_livres': colis_livres,
        'colis_avec_jc': colis_avec_jc,
        'title': f'Paiement du Lot {lot.numero_lot}'
    }
    
    return render(request, 'agent_mali_app/paiement_lot.html', context)

@agent_mali_required
def marquer_paiement_view(request, colis_id):
    """
    Marquer un colis comme payé et gérer les ajustements de type Jeton Cédé
    """
    if request.method == 'POST':
        colis = get_object_or_404(Colis, id=colis_id, statut='livre')
        
        # Mettre à jour toutes les livraisons de ce colis
        livraisons = colis.livraisons.filter(
            statut='livree',
            statut_paiement='en_attente'
        )
        
        if livraisons.exists():
            with transaction.atomic():
                # Appliquer les ajustements de type Jeton Cédé s'ils existent
                jc_adjustments = colis.price_adjustments.filter(
                    adjustment_type='jc',
                    status='active'
                )
                
                # Appliquer chaque ajustement Jeton Cédé
                for adjustment in jc_adjustments:
                    try:
                        adjustment.apply_adjustment()
                        messages.info(
                            request,
                            f"ℹ️ Ajustement Jeton Cédé de {adjustment.adjustment_amount} FCFA appliqué."
                        )
                    except Exception as e:
                        logger.error(f"Erreur lors de l'application de l'ajustement {adjustment.id}: {str(e)}")
                
                # Mettre à jour le statut de paiement des livraisons
                livraisons.update(statut_paiement='paye')
                
                # Mettre à jour le statut du colis si nécessaire
                if not colis.livraisons.filter(statut_paiement='en_attente').exists():
                    colis.statut = 'paye'
                    colis.save()
                
                messages.success(
                    request, 
                    f"✅ Colis {colis.numero_suivi} marqué comme payé."
                )
        else:
            messages.warning(
                request,
                f"⚠️ Aucune livraison en attente de paiement pour ce colis."
            )
    
    return redirect('agent_mali:colis_attente_paiement')

@agent_mali_required
def marquer_perdu_view(request, colis_id):
    """
    Marquer un colis comme perdu
    """
    if request.method == 'POST':
        colis = get_object_or_404(Colis, id=colis_id, statut='arrive')
        
        try:
            raison_perte = request.POST.get('raison_perte')
            commentaire = request.POST.get('commentaire', '')
            notifier_client = request.POST.get('notifier_client') == 'on'
            
            # Mettre à jour le statut du colis
            colis.statut = 'perdu'
            colis.save()
            
            # Envoyer notification au client si demandé
            if notifier_client:
                try:
                    from django.conf import settings
                    
                    message = f"""
😔 Information importante concernant votre colis

Bonjour {colis.client.user.get_full_name()},

Nous vous informons malheureusement que votre colis {colis.numero_suivi} a été marqué comme perdu.

📋 Raison: {dict([
    ('transport', 'Perdu pendant le transport'),
    ('vol', 'Vol/Disparition'),
    ('deterioration', 'Détérioration complète'),
    ('erreur_livraison', 'Erreur de livraison'),
    ('autre', 'Autre raison')
]).get(raison_perte, raison_perte)}

💬 Détails: {commentaire if commentaire else 'Aucun détail supplémentaire'}

Nous nous excusons sincèrement pour ce désagrément. Notre équipe va enquêter sur cet incident.

Veuillez nous contacter pour discuter des démarches de compensation.

Équipe TS Air Cargo Mali
📞 Contact: +223 XX XX XX XX
                    """.strip()
                    
                    NotificationService.send_notification(
                        user=colis.client.user,
                        message=message,
                        method='whatsapp',
                        title="Colis Perdu",
                        categorie='colis_perdu'
                    )
                    
                except Exception as notif_error:
                    # Éviter les prints en production; message d'erreur via messages ou logs si nécessaire
                    pass
            
            messages.success(
                request, 
                f"⚠️ Colis {colis.numero_suivi} marqué comme perdu. Client notifié: {'Oui' if notifier_client else 'Non'}"
            )
            
        except Exception as e:
            messages.error(request, f"❌ Erreur lors du marquage comme perdu: {str(e)}")
    
    return redirect('agent_mali:colis_a_livrer')

@agent_mali_required
@require_http_methods(["GET"])
def colis_details_api(request, colis_id):
    """
    API pour récupérer les détails d'un colis
    """
    try:
        colis = get_object_or_404(Colis, id=colis_id)
        
        # Informations du colis
        data = {
            'numero_suivi': colis.numero_suivi,
            'description': colis.description,
            'statut': colis.get_statut_display(),
            'type_transport': colis.get_type_transport_display(),
            'poids': float(colis.poids),
            'prix_calcule': float(colis.prix_calcule),
            'date_creation': colis.date_creation.strftime('%d/%m/%Y à %H:%M'),
            'dimensions': {
                'longueur': float(colis.longueur),
                'largeur': float(colis.largeur),
                'hauteur': float(colis.hauteur),
                'volume_m3': colis.volume_m3()
            },
            'client': {
                'nom': colis.client.user.get_full_name(),
                'telephone': colis.client.user.telephone,
                'email': colis.client.user.email or 'Non renseigné',
                'adresse': colis.client.adresse,
                'pays': colis.client.get_pays_display()
            },
            'lot': {
                'numero': colis.lot.numero_lot,
                'statut': colis.lot.get_statut_display(),
                'date_expedition': colis.lot.date_expedition.strftime('%d/%m/%Y à %H:%M') if colis.lot.date_expedition else None,
                'date_arrivee': colis.lot.date_arrivee.strftime('%d/%m/%Y à %H:%M') if colis.lot.date_arrivee else None,
                'agent_createur': colis.lot.agent_createur.get_full_name() if colis.lot.agent_createur else None
            }
        }
        
        # Informations de livraison si disponibles
        livraison = colis.livraisons.filter(statut='livree').first()
        if livraison:
            # Gérer le statut de paiement en toute sécurité
            try:
                if hasattr(livraison, 'statut_paiement') and livraison.statut_paiement:
                    statut_paiement = livraison.get_statut_paiement_display()
                else:
                    statut_paiement = 'Non défini'
            except:
                statut_paiement = 'Non défini'
            
            data['livraison'] = {
                'date_livraison': livraison.date_livraison_effective.strftime('%d/%m/%Y à %H:%M') if livraison.date_livraison_effective else 'Non définie',
                'nom_destinataire': livraison.nom_destinataire or 'Non défini',
                'adresse_livraison': livraison.adresse_livraison or 'Non définie',
                'statut_paiement': statut_paiement,
                'agent_livreur': livraison.agent_livreur.get_full_name() if livraison.agent_livreur else 'Non défini',
                'notes': livraison.notes_livraison or ''
            }
        
        return JsonResponse({
            'success': True,
            'data': data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@agent_mali_required
def lots_en_transit_view(request):
    """
    Liste des lots en transit venant de Chine (inclut les lots expédiés)
    """
    lots = Lot.objects.filter(
        statut__in=['expedie', 'en_transit']
    ).select_related('agent_createur').prefetch_related('colis__client__user').order_by('-date_expedition')
    
    # Calculs pour statistiques
    total_colis = sum(lot.colis.count() for lot in lots)
    valeur_transport_total = sum(float(lot.prix_transport or 0) for lot in lots)
    
    # Calcul de la valeur totale des colis (somme des prix des colis dans chaque lot)
    valeur_totale_colis = 0
    for lot in lots:
        valeur_totale_colis += sum(float(colis.prix_calcule or 0) for colis in lot.colis.all())
    
    context = {
        'lots': lots,
        'total_colis': total_colis,
        'valeur_transport_total': valeur_transport_total,
        'valeur_totale_colis': valeur_totale_colis,
        'title': 'Lots En Transit',
    }
    return render(request, 'agent_mali_app/lots_en_transit.html', context)

@agent_mali_required
def recevoir_lot_view(request, lot_id):
    """
    Réceptionner un lot au Mali avec gestion de la réception partielle
    """
    lot = get_object_or_404(Lot, id=lot_id, statut__in=['expedie', 'en_transit'])
    
    # Calculer la valeur totale des colis
    total_parcel_value = sum(colis.get_prix_effectif() for colis in lot.colis.all())
    
    if request.method == 'POST':
        try:
            commentaire = request.POST.get('commentaire', '')
            notifier_clients = request.POST.get('notifier_clients') == 'on'
            
            # Récupérer les frais de dédouanement
            frais_dedouanement = request.POST.get('frais_dedouanement', '0')
            try:
                frais_dedouanement = float(frais_dedouanement) if frais_dedouanement else 0.0
                if frais_dedouanement < 0:
                    frais_dedouanement = 0.0
            except (ValueError, TypeError):
                frais_dedouanement = 0.0
            
            # Récupérer les colis sélectionnés pour réception
            colis_recus_ids = request.POST.getlist('colis_recus')
            
            # Si aucun colis spécifique n'est sélectionné, 
            # réceptionner seulement les colis pas encore arrivés
            if not colis_recus_ids:
                colis_non_arrives = lot.colis.exclude(statut='arrive')
                colis_recus_ids = [str(c.id) for c in colis_non_arrives]
            
            # S'assurer qu'on ne traite que les colis pas encore arrivés
            colis_a_recevoir = lot.colis.filter(
                id__in=colis_recus_ids
            ).exclude(statut='arrive')
            
            if not colis_a_recevoir.exists():
                messages.warning(request, "⚠️ Aucun colis nouveau à réceptionner dans cette sélection.")
                return redirect('agent_mali:recevoir_lot', lot_id=lot.id)
            
            # Créer ou récupérer l'enregistrement de réception
            reception, created = ReceptionLot.objects.get_or_create(
                lot=lot,
                defaults={
                    'agent_receptionnaire': request.user,
                    'reception_complete': False,
                    'nombre_colis_recus': 0,
                    'frais_dedouanement': frais_dedouanement
                }
            )
            
            # Mettre à jour les frais de dédouanement si ce n'est pas une nouvelle réception
            if not created and frais_dedouanement > 0:
                reception.frais_dedouanement = frais_dedouanement
                
            # Mettre à jour les frais de douane du lot (utilisé pour le calcul du bénéfice)
            # Mettre à jour même si frais_dedouanement est à 0
            lot.frais_douane = frais_dedouanement
            lot.save(update_fields=['frais_douane'])  # Sauvegarder explicitement le champ frais_douane
            
            # Ajouter l'observation avec horodatage
            if commentaire:
                reception.ajouter_observation(commentaire)
            else:
                reception.ajouter_observation("Réception effectuée sans commentaire")
            
            # Mettre à jour le statut des colis reçus
            colis_a_recevoir.update(statut='arrive')
            
            # Mettre à jour le nombre de colis reçus
            reception.nombre_colis_recus = lot.colis.filter(statut='arrive').count()
            
            # Vérifier si tous les colis du lot sont maintenant arrivés
            colis_non_arrives_restants = lot.colis.exclude(statut='arrive')
            total_colis = lot.colis.count()
            
            if not colis_non_arrives_restants.exists():
                # Réception complète
                lot.statut = 'arrive'
                reception.reception_complete = True
                reception.colis_manquants.clear()
                reception_type = "complète"
                reception_action = "Première et dernière" if created else "Dernière"
                
                # Ajouter une observation de réception complète
                reception.ajouter_observation(
                    f"RÉCEPTION COMPLÈTE - {reception.nombre_colis_recus}/{total_colis} colis reçus. "
                    f"Tous les colis ont été réceptionnés avec succès."
                )
            else:
                # Réception partielle
                lot.statut = 'en_transit'
                reception.reception_complete = False
                reception.colis_manquants.set(colis_non_arrives_restants)
                reception_type = "partielle"
                reception_action = "Première" if created else "Nouvelle"
                
                # Ajouter une observation de réception partielle
                reception.ajouter_observation(
                    f"RÉCEPTION PARTIELLE - {reception.nombre_colis_recus}/{total_colis} colis reçus. "
                    f"Colis manquants: {colis_non_arrives_restants.count()}"
                )
            
            # Mettre à jour les dates
            lot.date_arrivee = timezone.now()
            if created:
                reception.date_reception = timezone.now()
            
            # Sauvegarder les modifications
            reception.save()
            # Le lot est déjà sauvegardé avec les frais de douane, on le sauvegarde à nouveau pour le bénéfice
            lot.save()
            
            # Envoyer des notifications de masse via tâche asynchrone
            colis_recus_count = colis_a_recevoir.count()
            
            # Logs debug supprimés pour la production
            
            if notifier_clients:
                # Utiliser la nouvelle tâche qui ne notifie que les colis réellement réceptionnés
                from notifications_app.tasks import send_bulk_received_colis_notifications
                
                # Récupérer les IDs des colis qui viennent d'être réceptionnés
                colis_recus_ids_int = [int(colis_id) for colis_id in colis_recus_ids]
                
                try:
                    # Lancer la tâche asynchrone pour les colis réceptionnés uniquement
                    task_result = send_bulk_received_colis_notifications.delay(
                        colis_ids_list=colis_recus_ids_int,
                        notification_type='lot_arrived',
                        initiated_by_id=request.user.id
                    )
                    
                    notifications_envoyees = f"Tâche asynchrone lancée (ID: {task_result.id})"
                    
                except Exception as async_error:
                    # En production on évite les prints; laisser remonter via messages/monitoring
                    pass
                    notifications_envoyees = "Erreur lors du lancement"
            
            if notifier_clients:
                if "Tâche asynchrone" in str(notifications_envoyees):
                    messages.success(request, f"✅ {reception_action} réception {reception_type} du lot {lot.numero_lot} ! {colis_recus_count} colis reçus. Notifications en cours d'envoi...")
                else:
                    messages.success(request, f"✅ {reception_action} réception {reception_type} du lot {lot.numero_lot} ! {colis_recus_count} colis reçus. Erreur notifications: {notifications_envoyees}")
            else:
                messages.success(request, f"✅ {reception_action} réception {reception_type} du lot {lot.numero_lot} ! {colis_recus_count} colis reçus (notifications désactivées).")
            return redirect('agent_mali:lots_en_transit')
            
        except Exception as e:
            messages.error(request, f"❌ Erreur lors de la réception: {str(e)}")
    
    # Calculs pour affichage
    total_colis = lot.colis.count()
    total_poids = sum(float(c.poids) for c in lot.colis.all())
    total_parcel_value = sum(colis.get_prix_effectif() for colis in lot.colis.all())
    
    # Vérifier si c'est la première réception du lot
    is_first_reception = True
    if hasattr(lot, 'reception_mali'):
        # C'est la première réception si aucun colis n'a encore été reçu
        is_first_reception = lot.reception_mali.nombre_colis_recus == 0
    
    context = {
        'lot': lot,
        'total_colis': total_colis,
        'total_poids': total_poids,
        'total_parcel_value': total_parcel_value,
        'is_first_reception': is_first_reception,
        'title': f'Réceptionner le lot {lot.numero_lot}',
    }
    return render(request, 'agent_mali_app/recevoir_lot.html', context)

@agent_mali_required
def colis_a_livrer_view(request):
    """
    Liste des colis arrivés au Mali et prêts à être livrés
    """
    colis = Colis.objects.filter(
        statut='arrive'
    ).select_related('client__user', 'lot').order_by('-lot__date_arrivee')
    
    # Filtrage par recherche
    search_query = request.GET.get('search', '')
    if search_query:
        colis = colis.filter(
            Q(numero_suivi__icontains=search_query) |
            Q(client__user__first_name__icontains=search_query) |
            Q(client__user__last_name__icontains=search_query) |
            Q(client__user__telephone__icontains=search_query)
        )
    
    # Calcul de la valeur totale
    valeur_totale = sum(float(c.prix_calcule) for c in colis)
    
    # Calcul du poids total
    poids_total = sum(float(c.poids) for c in colis)
    
    # Calcul des clients uniques
    clients_uniques = colis.values_list('client__id', flat=True).distinct().count()
    
    context = {
        'colis': colis,
        'search_query': search_query,
        'valeur_totale': valeur_totale,
        'poids_total': poids_total,
        'clients_uniques': clients_uniques,
        'title': 'Colis à Livrer',
    }
    return render(request, 'agent_mali_app/colis_a_livrer.html', context)

@agent_mali_required
def marquer_livre_view(request, colis_id):
    """
    Marquer un colis comme livré
    """
    colis = get_object_or_404(Colis, id=colis_id, statut='arrive')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            nom_destinataire = request.POST.get('personne_receptrice') or colis.client.user.get_full_name()
            adresse_livraison = request.POST.get('adresse_livraison') or colis.client.adresse or "Adresse non spécifiée"
            notes_livraison = request.POST.get('commentaire', '')
            mode_livraison = request.POST.get('mode_livraison')
            statut_paiement = request.POST.get('statut_paiement', 'en_attente')
            
            # Gestion du jeton cédé (JC) si fourni
            jeton_cede_str = request.POST.get('jeton_cede', '0')
            jeton_cede = 0.0
            
            try:
                jeton_cede = float(jeton_cede_str) if jeton_cede_str else 0.0
                
                if jeton_cede > 0:
                    # Vérifier que le jeton ne dépasse pas 50% du prix du colis
                    prix_effectif = float(colis.get_prix_effectif())
                    if jeton_cede > prix_effectif * 0.5:
                        messages.error(request, "Le jeton cédé ne peut pas dépasser 50% du prix du colis.")
                        return redirect('agent_mali:marquer_livre', colis_id=colis_id)
                    
                    # Créer l'ajustement de prix pour le jeton cédé
                    prix_final = prix_effectif - jeton_cede
                    
                    PriceAdjustment.objects.create(
                        colis=colis,
                        adjustment_type='jc',
                        adjustment_amount=jeton_cede,
                        original_price=prix_effectif,
                        final_price=prix_final,
                        reason="Jeton cédé lors de la livraison",
                        notes=f"Jeton cédé par {request.user.get_full_name()} le {timezone.now().strftime('%d/%m/%Y %H:%M')}",
                        status='applied',
                        applied_by=request.user,
                        created_at=timezone.now()
                    )
                    
                    # Mettre à jour le statut de paiement si le jeton est supérieur ou égal au prix
                    if jeton_cede >= prix_effectif:
                        statut_paiement = 'paye'
                        notes_livraison = f"{notes_livraison}\n\nLe jeton cédé couvre l'intégralité du montant dû.".strip()
            except (ValueError, TypeError) as e:
                messages.error(request, f"Erreur de format du jeton cédé: {str(e)}")
                return redirect('agent_mali:marquer_livre', colis_id=colis_id)
            
            # Créer l'enregistrement de livraison
            livraison = Livraison.objects.create(
                colis=colis,
                agent_livreur=request.user,
                date_planifiee=timezone.now(),
                date_livraison_effective=timezone.now(),
                statut='livree',
                statut_paiement=statut_paiement,
                adresse_livraison=adresse_livraison,
                telephone_destinataire=colis.client.user.telephone,
                nom_destinataire=nom_destinataire,
                notes_livraison=notes_livraison
            )
            
            # Mettre à jour le statut du colis et le prix final
            colis.statut = 'livre'
            colis.save()
            
            # Mettre à jour le statut de paiement si le jeton couvre le montant total
            if jeton_cede and jeton_cede >= colis.get_prix_effectif():
                livraison.statut_paiement = 'paye'
                livraison.save()
            
            # Ajouter une note sur le jeton cédé si applicable
            if jeton_cede and float(jeton_cede) > 0:
                notes_livraison = f"{notes_livraison}\n\nJeton cédé: {jeton_cede} FCFA".strip()
                livraison.notes_livraison = notes_livraison
                livraison.save()
            
            # Envoyer notification de livraison
            try:
                from django.conf import settings
                
                if getattr(settings, 'DEBUG', True):
                    message = f"""
✅ [SANDBOX TEST] Colis livré avec succès !

👤 Client: {colis.client.user.get_full_name()}
📞 Téléphone: {colis.client.user.telephone}

📦 Colis: {colis.numero_suivi}
🏠 Livré à: {nom_destinataire}
📅 Date de livraison: {livraison.date_livraison_effective.strftime('%d/%m/%Y à %H:%M')}

✅ Votre colis a été livré avec succès !
Merci d'avoir choisi TS Air Cargo.

Équipe TS Air Cargo 🚀
                    """.strip()
                else:
                    message = f"""
✅ Colis livré avec succès !

Votre colis {colis.numero_suivi} a été livré avec succès.

📅 Date de livraison: {livraison.date_livraison_effective.strftime('%d/%m/%Y à %H:%M')}
🏠 Livré à: {nom_destinataire}

Merci d'avoir choisi TS Air Cargo !
                    """.strip()
                
                NotificationService.send_notification(
                    user=colis.client.user,
                    message=message,
                    method='whatsapp',
                    title="Colis Livré",
                    categorie='colis_livre'
                )
            except Exception as notif_error:
                # Éviter les prints en production
                pass
            
            messages.success(request, f"✅ Colis {colis.numero_suivi} marqué comme livré avec succès !")
            return redirect('agent_mali:colis_a_livrer')
            
        except Exception as e:
            messages.error(request, f"❌ Erreur lors de la livraison: {str(e)}")
    
    context = {
        'colis': colis,
        'title': f'Livrer le colis {colis.numero_suivi}',
    }
    return render(request, 'agent_mali_app/marquer_livre.html', context)

@agent_mali_required
def depenses_view(request):
    """
    Liste et gestion des dépenses avec statistiques avancées
    """
    # Base queryset
    depenses = Depense.objects.filter(agent=request.user).order_by('-date_depense')
    
    # Filtres
    search_query = request.GET.get('search', '')
    categorie_filter = request.GET.get('categorie', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    montant_min = request.GET.get('montant_min', '')
    
    # Application des filtres
    if search_query:
        depenses = depenses.filter(
            Q(libelle__icontains=search_query) |
            Q(notes__icontains=search_query) |
            Q(montant__icontains=search_query)
        )
    
    if categorie_filter:
        depenses = depenses.filter(type_depense=categorie_filter)
    
    if date_debut:
        depenses = depenses.filter(date_depense__gte=date_debut)
    
    if date_fin:
        depenses = depenses.filter(date_depense__lte=date_fin)
    
    if montant_min:
        try:
            depenses = depenses.filter(montant__gte=float(montant_min))
        except ValueError:
            pass
    
    # Calcul des statistiques
    aujourd_hui = date.today()
    debut_mois = aujourd_hui.replace(day=1)
    debut_semaine = aujourd_hui - timedelta(days=aujourd_hui.weekday())
    
    # Totaux par période
    total_mois = Depense.objects.filter(
        agent=request.user,
        date_depense__gte=debut_mois
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    total_semaine = Depense.objects.filter(
        agent=request.user,
        date_depense__gte=debut_semaine
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    # Moyenne journalière (30 derniers jours)
    debut_30j = aujourd_hui - timedelta(days=30)
    total_30j = Depense.objects.filter(
        agent=request.user,
        date_depense__gte=debut_30j
    ).aggregate(total=Sum('montant'))['total'] or 0
    moyenne_journaliere = total_30j / 30
    
    # Pagination
    paginator = Paginator(depenses, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'depenses': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'total_mois': float(total_mois),
        'total_semaine': float(total_semaine),
        'moyenne_journaliere': float(moyenne_journaliere),
        'search_query': search_query,
        'categorie_filter': categorie_filter,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'montant_min': montant_min,
        'title': 'Gestion des Dépenses',
    }
    return render(request, 'agent_mali_app/depenses.html', context)

@agent_mali_required
def appliquer_ajustement_view(request, colis_id):
    """
    Interface pour appliquer un ajustement de prix (JC ou remise) à un colis
    """
    colis = get_object_or_404(Colis, id=colis_id)
    
    # Vérifier que le colis peut être ajusté
    if colis.statut not in ['arrive', 'livre']:
        messages.error(request, f"❌ Impossible d'ajuster le prix d'un colis au statut '{colis.get_statut_display()}'")
        return redirect('agent_mali:colis_a_livrer')
    
    if request.method == 'POST':
        try:
            adjustment_type = request.POST.get('adjustment_type')
            amount = float(request.POST.get('amount', 0))
            reason = request.POST.get('reason', '').strip()
            notes = request.POST.get('notes', '').strip()
            
            # Validation
            if not adjustment_type or adjustment_type not in ['jc', 'remise']:
                messages.error(request, "❌ Type d'ajustement requis")
                return render(request, 'agent_mali_app/appliquer_ajustement.html', {'colis': colis})
            
            if amount <= 0:
                messages.error(request, "❌ Le montant doit être supérieur à 0")
                return render(request, 'agent_mali_app/appliquer_ajustement.html', {'colis': colis})
            
            if not reason:
                messages.error(request, "❌ La raison est requise")
                return render(request, 'agent_mali_app/appliquer_ajustement.html', {'colis': colis})
            
            # Créer l'ajustement
            adjustment = PriceAdjustment.objects.create(
                colis=colis,
                adjustment_type=adjustment_type,
                adjustment_amount=amount,
                original_price=colis.get_prix_effectif(),
                reason=reason,
                notes=notes,
                applied_by=request.user
            )
            
            # Appliquer immédiatement l'ajustement
            adjustment.apply_adjustment()
            
            type_label = "Jeton cédé" if adjustment_type == 'jc' else "Remise"
            messages.success(
                request, 
                f"✅ {type_label} de {amount} FCFA appliqué au colis {colis.numero_suivi}. "
                f"Prix passé de {adjustment.original_price} à {adjustment.final_price} FCFA."
            )
            
            return redirect('agent_mali:colis_detail', colis_id=colis.id)
            
        except ValueError as e:
            messages.error(request, f"❌ Erreur dans les données saisies: {str(e)}")
        except Exception as e:
            messages.error(request, f"❌ Erreur lors de l'application de l'ajustement: {str(e)}")
    
    # Calculer les suggestions de jetons communs
    prix_actuel = colis.get_prix_effectif()
    suggestions_jc = []
    for jeton in [25, 50, 100, 200, 250, 500]:
        if jeton < prix_actuel:
            suggestions_jc.append(jeton)
    
    context = {
        'colis': colis,
        'prix_actuel': prix_actuel,
        'suggestions_jc': suggestions_jc,
        'title': f'Ajuster le prix du colis {colis.numero_suivi}',
    }
    return render(request, 'agent_mali_app/appliquer_ajustement.html', context)

@agent_mali_required
def colis_detail_view(request, colis_id):
    """
    Détail d'un colis avec historique des ajustements
    """
    colis = get_object_or_404(Colis, id=colis_id)
    
    # Historique des ajustements
    ajustements = colis.price_adjustments.all().order_by('-created_at')
    
    # Calculs financiers
    prix_original = colis.prix_calcule
    prix_actuel = colis.get_prix_effectif()
    total_ajustements = sum(adj.effective_adjustment for adj in ajustements.filter(status='applied'))
    
    context = {
        'colis': colis,
        'ajustements': ajustements,
        'prix_original': prix_original,
        'prix_actuel': prix_actuel,
        'total_ajustements': total_ajustements,
        'title': f'Détail du colis {colis.numero_suivi}',
    }
    return render(request, 'agent_mali_app/colis_detail.html', context)

@agent_mali_required
def annuler_ajustement_view(request, adjustment_id):
    """
    Annuler un ajustement de prix
    """
    if request.method == 'POST':
        adjustment = get_object_or_404(
            PriceAdjustment, 
            id=adjustment_id,
            applied_by=request.user  # Seul l'agent qui l'a créé peut l'annuler
        )
        
        try:
            colis_id = adjustment.colis.id
            type_label = adjustment.get_adjustment_type_display()
            
            adjustment.cancel_adjustment()
            
            messages.success(
                request,
                f"✅ {type_label} de {adjustment.adjustment_amount} FCFA annulé pour le colis {adjustment.colis.numero_suivi}"
            )
            
            return redirect('agent_mali:colis_detail', colis_id=colis_id)
            
        except Exception as e:
            messages.error(request, f"❌ Erreur lors de l'annulation: {str(e)}")
    
    return redirect('agent_mali:colis_a_livrer')

@agent_mali_required
def ajustements_rapport_view(request):
    """
    Rapport détaillé des ajustements de prix
    """
    # Filtres de date
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    type_filter = request.GET.get('type', '')
    
    # Base queryset
    ajustements = PriceAdjustment.objects.filter(
        status='applied'
    ).select_related('colis__client__user', 'applied_by').order_by('-created_at')
    
    # Filtres
    if date_debut:
        try:
            date_debut_parsed = datetime.strptime(date_debut, '%Y-%m-%d').date()
            ajustements = ajustements.filter(created_at__date__gte=date_debut_parsed)
        except ValueError:
            pass
    
    if date_fin:
        try:
            date_fin_parsed = datetime.strptime(date_fin, '%Y-%m-%d').date()
            ajustements = ajustements.filter(created_at__date__lte=date_fin_parsed)
        except ValueError:
            pass
    
    if type_filter:
        ajustements = ajustements.filter(adjustment_type=type_filter)
    
    # Statistiques
    total_ajustements = ajustements.count()
    total_jetons_cedes = ajustements.filter(adjustment_type='jc').aggregate(total=Sum('adjustment_amount'))['total'] or 0
    total_remises = ajustements.filter(adjustment_type='remise').aggregate(total=Sum('adjustment_amount'))['total'] or 0
    
    # Pagination
    paginator = Paginator(ajustements, 25)
    page_number = request.GET.get('page')
    ajustements_page = paginator.get_page(page_number)
    
    context = {
        'ajustements': ajustements_page,
        'total_ajustements': total_ajustements,
        'total_jetons_cedes': total_jetons_cedes,
        'total_remises': total_remises,
        'total_montant': total_jetons_cedes + total_remises,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'type_filter': type_filter,
        'type_choices': PriceAdjustment.ADJUSTMENT_TYPES,
        'title': 'Rapport des Ajustements de Prix',
    }
    return render(request, 'agent_mali_app/ajustements_rapport.html', context)

@agent_mali_required
def depense_create_view(request):
    """
    Créer une nouvelle dépense
    """
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            libelle = request.POST.get('libelle')
            type_depense = request.POST.get('type_depense')
            montant = request.POST.get('montant')
            date_depense = request.POST.get('date_depense')
            notes = request.POST.get('notes', '')
            justificatif = request.FILES.get('justificatif')
            
            # Créer la dépense
            depense = Depense.objects.create(
                libelle=libelle,
                type_depense=type_depense,
                montant=float(montant),
                date_depense=date_depense,
                agent=request.user,
                notes=notes,
                justificatif=justificatif
            )
            
            messages.success(request, f"✅ Dépense '{libelle}' enregistrée avec succès !")
            return redirect('agent_mali:depenses')
            
        except Exception as e:
            messages.error(request, f"❌ Erreur lors de l'enregistrement: {str(e)}")
    
    # Statistiques pour le sidebar
    aujourd_hui = date.today()
    debut_mois = aujourd_hui.replace(day=1)
    
    total_mois_actuel = Depense.objects.filter(
        agent=request.user,
        date_depense__gte=debut_mois
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    nombre_depenses_mois = Depense.objects.filter(
        agent=request.user,
        date_depense__gte=debut_mois
    ).count()
    
    moyenne_depense_mois = total_mois_actuel / nombre_depenses_mois if nombre_depenses_mois > 0 else 0
    
    dernieres_depenses = Depense.objects.filter(
        agent=request.user
    ).order_by('-date_depense')[:5]
    
    context = {
        'title': 'Nouvelle Dépense',
        'type_choices': Depense.TYPE_DEPENSE_CHOICES,
        'total_mois_actuel': float(total_mois_actuel),
        'nombre_depenses_mois': nombre_depenses_mois,
        'moyenne_depense_mois': float(moyenne_depense_mois),
        'dernieres_depenses': dernieres_depenses,
    }
    return render(request, 'agent_mali_app/nouvelle_depense.html', context)

@agent_mali_required
def nouvelle_depense_view(request):
    """
    Créer une nouvelle dépense - alias pour depense_create_view
    """
    return depense_create_view(request)


@agent_mali_required
def depense_edit_view(request, depense_id):
    """
    Modifier une dépense existante
    """
    depense = get_object_or_404(Depense, id=depense_id, agent=request.user)
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            libelle = request.POST.get('libelle')
            type_depense = request.POST.get('type_depense')
            montant = request.POST.get('montant')
            date_depense = request.POST.get('date_depense')
            notes = request.POST.get('notes', '')
            justificatif = request.FILES.get('justificatif')
            
            # Mettre à jour la dépense
            depense.libelle = libelle
            depense.type_depense = type_depense
            depense.montant = float(montant)
            depense.date_depense = date_depense
            depense.notes = notes
            
            # Mettre à jour le justificatif si fourni
            if justificatif:
                depense.justificatif = justificatif
            
            depense.save()
            
            messages.success(request, f"✅ Dépense '{libelle}' modifiée avec succès !")
            return redirect('agent_mali:depenses')
            
        except Exception as e:
            messages.error(request, f"❌ Erreur lors de la modification: {str(e)}")
    
    context = {
        'depense': depense,
        'title': f'Modifier Dépense - {depense.libelle}',
        'type_choices': Depense.TYPE_DEPENSE_CHOICES,
    }
    return render(request, 'agent_mali_app/depense_form.html', context)

@agent_mali_required
def depense_delete_view(request, depense_id):
    """
    Supprimer une dépense
    """
    depense = get_object_or_404(Depense, id=depense_id, agent=request.user)
    
    if request.method == 'POST':
        libelle = depense.libelle
        depense.delete()
        messages.success(request, f"✅ Dépense '{libelle}' supprimée avec succès !")
        return redirect('agent_mali:depenses')
    
    return redirect('agent_mali:depenses')

@agent_mali_required
def depense_detail_view(request, depense_id):
    """
    Afficher les détails d'une dépense
    """
    depense = get_object_or_404(Depense, id=depense_id, agent=request.user)
    
    context = {
        'depense': depense,
        'title': f'Détails - {depense.libelle}',
    }
    return render(request, 'agent_mali_app/depense_detail.html', context)

@agent_mali_required
def rapports_view(request):
    """
    Interface de génération de rapports
    """
    # Statistiques du jour pour affichage
    aujourd_hui = date.today()
    
    # Lots reçus aujourd'hui
    lots_recus_aujourd_hui = ReceptionLot.objects.filter(
        date_reception__date=aujourd_hui
    ).count()
    
    # Colis livrés aujourd'hui
    colis_livres_aujourd_hui = Livraison.objects.filter(
        date_livraison_effective__date=aujourd_hui,
        statut='livree'
    ).count()
    
    # Dépenses du jour
    depenses_total_aujourd_hui = Depense.objects.filter(
        agent=request.user,
        date_depense=aujourd_hui
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    # Clients notifiés aujourd'hui (approximation)
    clients_notifies_aujourd_hui = lots_recus_aujourd_hui + colis_livres_aujourd_hui
    
    # Revenus de livraison aujourd'hui
    revenus_livraison_aujourd_hui = Livraison.objects.filter(
        date_livraison_effective__date=aujourd_hui,
        statut='livree',
        montant_collecte__isnull=False
    ).aggregate(total=Sum('montant_collecte'))['total'] or 0
    
    # Historique des rapports (simulé pour l'instant)
    rapports_historique = []
    
    context = {
        'stats_jour': {
            'lots_recus': lots_recus_aujourd_hui,
            'colis_livres': colis_livres_aujourd_hui,
            'depenses_total': float(depenses_total_aujourd_hui),
            'clients_notifies': clients_notifies_aujourd_hui,
            'revenus_livraison': float(revenus_livraison_aujourd_hui),
        },
        'rapports_historique': rapports_historique,
        'today': aujourd_hui,
        'title': 'Rapports et Analyses',
    }
    return render(request, 'agent_mali_app/rapports.html', context)

@agent_mali_required
@require_http_methods(["POST"])
def send_report_whatsapp_api(request):
    """
    API pour envoyer un rapport via WhatsApp ET par email avec PDF
    """
    try:
        data = json.loads(request.body)
        report_type = data.get('type')
        period = data.get('period')
        
        # Générer le rapport PDF
        pdf_content = generate_daily_report_pdf(period)
        
        # Message de base pour WhatsApp
        whatsapp_message = f"""
📈 Rapport Journalier TS Air Cargo Mali

📅 Date: {datetime.strptime(period, '%Y-%m-%d').strftime('%d/%m/%Y')}
👥 Agent: {request.user.get_full_name()}
🏢 Agence: Mali

📊 Rapport automatique quotidien généré.

📧 Le rapport détaillé en PDF a été envoyé par email.

Équipe TS Air Cargo Mali 🚀
        """.strip()
        
        # Utiliser le service de notification existant pour simplifier
        from notifications_app.services import NotificationService
        
        # Créer un message détaillé pour les admins WhatsApp
        admin_whatsapp_message = f"""
📈 RAPPORT JOURNALIER AUTOMATIQUE

📅 Date: {datetime.strptime(period, '%Y-%m-%d').strftime('%d/%m/%Y')}
👥 Généré par: {request.user.get_full_name()}
🏢 Agence: Mali - Bamako

📆 Type: Rapport {report_type}
🗓️ Période: {period}

📧 Le rapport détaillé en PDF a été envoyé par email.

Équipe TS Air Cargo Mali 🚀
        """.strip()
        
        # 1. Envoyer la notification WhatsApp
        admins_contacted_whatsapp = NotificationService.send_admin_notification(
            message=admin_whatsapp_message,
            title=f"Rapport {report_type} - {datetime.strptime(period, '%Y-%m-%d').strftime('%d/%m/%Y')}"
        )
        
        # 2. Envoyer le PDF par email
        from django.core.mail import EmailMessage
        from django.conf import settings
        
        # Sujet de l'email
        email_subject = f"📊 Rapport Quotidien TS Air Cargo Mali - {datetime.strptime(period, '%Y-%m-%d').strftime('%d/%m/%Y')}"
        
        # Corps de l'email
        email_body = f"""
Bonjour,

Veuillez trouver ci-joint le rapport quotidien d'activité TS Air Cargo Mali.

📊 INFORMATIONS DU RAPPORT:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
📅 Date du rapport: {datetime.strptime(period, '%Y-%m-%d').strftime('%d/%m/%Y')}
👤 Généré par: {request.user.get_full_name()}
🏢 Agence: Mali - Bamako
🕐 Date de génération: {datetime.now().strftime('%d/%m/%Y à %H:%M')}

📋 CONTENU DU RAPPORT:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• 📦 Lots reçus de Chine
• 🚚 Colis livrés aux clients
• 💸 Dépenses enregistrées
• 💰 Revenus de livraison collectés
• 📊 Analyses de performance journalière
• 📈 Indicateurs clés de performance (KPI)

📎 Le rapport détaillé est disponible en pièce jointe au format PDF.

📱 NOTIFICATION WHATSAPP:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Une notification WhatsApp a également été envoyée à {admins_contacted_whatsapp} administrateur(s).

Pour toute question concernant ce rapport, n'hésitez pas à nous contacter.

Cordialement,
Équipe TS Air Cargo Mali

📧 contact@ts-aircargo.com | 📞 +223 XX XX XX XX
🌐 www.ts-aircargo.com
        """.strip()
        
        # Déterminer les destinataires de l'email
        email_recipients = getattr(settings, 'EMAIL_REPORT_RECIPIENTS', ['admin@ts-aircargo.com'])
        
        # En mode développement, vérifier si l'utilisateur a un email
        if settings.DEBUG and request.user.email:
            # Ajouter l'email de l'utilisateur connecté pour les tests
            if request.user.email not in email_recipients:
                email_recipients = [request.user.email] + email_recipients
        
        # Créer et envoyer l'email
        email_sent_count = 0
        try:
            email = EmailMessage(
                subject=email_subject,
                body=email_body,
                from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@ts-aircargo.com'),
                to=email_recipients,
            )
            
            # Attacher le PDF
            email.attach(
                f"rapport_quotidien_{period}.pdf",
                pdf_content,
                'application/pdf'
            )
            
            # Envoyer l'email
            email.send()
            email_sent_count = len(email_recipients)
            
        except Exception as email_error:
            # Éviter les prints en production; continuer même si l'email échoue
            pass
        
        # Construire le message de retour
        success_parts = []
        if admins_contacted_whatsapp > 0:
            success_parts.append(f"{admins_contacted_whatsapp} notification(s) WhatsApp")
        if email_sent_count > 0:
            success_parts.append(f"{email_sent_count} email(s) avec PDF")
        
        if success_parts:
            return JsonResponse({
                'success': True,
                'message': f'Rapport {report_type} envoyé avec succès : ' + ' et '.join(success_parts)
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'Aucune notification n\'a pu être envoyée. Vérifiez la configuration WhatsApp et Email.'
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@agent_mali_required
@require_http_methods(["POST"])
def generate_daily_report_api(request):
    """
    API pour générer un rapport quotidien en PDF et le renvoyer comme un fichier
    """
    try:
        data = json.loads(request.body)
        date = data.get('date')

        # Vérifier la date
        if not date:
            return JsonResponse({'success': False, 'error': 'Date non fournie'}, status=400)

        # Générer le PDF
        pdf_content = generate_daily_report_pdf(date)

        # Retourner le fichier PDF
        response = HttpResponse(pdf_content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="rapport_quotidien_{date}.pdf"'
        return response

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@agent_mali_required
@require_http_methods(["POST"])
def generate_monthly_report_api(request):
    """
    API pour générer un rapport mensuel en PDF et le renvoyer comme un fichier
    """
    try:
        data = json.loads(request.body)
        month = data.get('month')  # Format: 'YYYY-MM'

        # Vérifier le mois
        if not month:
            return JsonResponse({'success': False, 'error': 'Mois non fourni'}, status=400)

        # Générer le PDF
        pdf_content = generate_monthly_report_pdf(month)

        # Retourner le fichier PDF
        response = HttpResponse(pdf_content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="rapport_mensuel_{month}.pdf"'
        return response

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@agent_mali_required
@require_http_methods(["POST"])
def generate_yearly_report_api(request):
    """
    API pour générer un rapport annuel en PDF et le renvoyer comme un fichier
    """
    try:
        data = json.loads(request.body)
        year = data.get('year')  # Format: 'YYYY'

        # Vérifier l'année
        if not year:
            return JsonResponse({'success': False, 'error': 'Année non fournie'}, status=400)

        # Générer le PDF
        pdf_content = generate_yearly_report_pdf(year)

        # Retourner le fichier PDF
        response = HttpResponse(pdf_content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="rapport_annuel_{year}.pdf"'
        return response

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@agent_mali_required
@require_http_methods(["POST"])
def send_report_email_api(request):
    """
    API pour envoyer un rapport par email
    """
    try:
        data = json.loads(request.body)
        report_type = data.get('type', 'daily')
        period = data.get('period')
        
        if not period:
            return JsonResponse({
                'success': False,
                'error': 'Période non fournie'
            })
        
        # Générer le PDF selon le type de rapport
        if report_type == 'daily':
            pdf_content = generate_daily_report_pdf(period)
            subject = f"Rapport Quotidien TS Air Cargo Mali - {datetime.strptime(period, '%Y-%m-%d').strftime('%d/%m/%Y')}"
            filename = f"rapport_quotidien_{period}.pdf"
        elif report_type == 'monthly':
            pdf_content = generate_monthly_report_pdf(period)
            year, month = period.split('-')
            month_name = datetime(int(year), int(month), 1).strftime('%B %Y')
            subject = f"Rapport Mensuel TS Air Cargo Mali - {month_name}"
            filename = f"rapport_mensuel_{period}.pdf"
        elif report_type == 'yearly':
            pdf_content = generate_yearly_report_pdf(period)
            subject = f"Rapport Annuel TS Air Cargo Mali - {period}"
            filename = f"rapport_annuel_{period}.pdf"
        else:
            return JsonResponse({
                'success': False,
                'error': 'Type de rapport non supporté'
            })
        
        # Contenu de l'email
        email_body = f"""
Bonjour,

Veuillez trouver ci-joint le rapport d'activité TS Air Cargo Mali.

📊 Type de rapport: {report_type.title()}
📅 Période: {period}
👤 Généré par: {request.user.get_full_name()}
🏢 Agence: Mali - Bamako
🕐 Date de génération: {datetime.now().strftime('%d/%m/%Y à %H:%M')}

Ce rapport contient les informations détaillées sur:
• Les lots reçus
• Les colis livrés
• Les dépenses enregistrées
• Les revenus générés
• Les analyses de performance

Cordialement,
Équipe TS Air Cargo Mali
        """.strip()
        
        # Envoyer l'email
        from django.core.mail import EmailMessage
        from django.conf import settings
        
        # Configuration email destinataire
        # En mode développement, envoyer à l'utilisateur connecté ou un email de test
        if hasattr(settings, 'EMAIL_TEST_RECIPIENTS') and settings.EMAIL_TEST_RECIPIENTS:
            recipients = settings.EMAIL_TEST_RECIPIENTS
        else:
            # Fallback: utiliser l'email de l'utilisateur connecté si disponible
            recipients = [request.user.email] if request.user.email else ['admin@ts-aircargo.com']
        
        email = EmailMessage(
            subject=subject,
            body=email_body,
            from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@ts-aircargo.com'),
            to=recipients,
        )
        
        # Attacher le PDF
        email.attach(filename, pdf_content, 'application/pdf')
        
        # Envoyer l'email
        email.send()
        
        return JsonResponse({
            'success': True,
            'message': f'Rapport {report_type} envoyé par email avec succès à {len(recipients)} destinataire(s)'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors de l\'envoi email: {str(e)}'
        })

@agent_mali_required
@require_http_methods(["POST"])
def schedule_auto_report_api(request):
    """
    API pour programmer l'envoi automatique des rapports
    """
    try:
        data = json.loads(request.body)
        enabled = data.get('enabled', False)
        time = data.get('time', '23:59')
        
        if enabled:
            # Programmer la tâche automatique
            # Note: Ceci nécessite une implémentation avec Celery ou Django-cron
            # Pour le moment, on simule la programmation
            
            # Sauvegarder la configuration dans la session ou la base de données
            request.session['auto_report_enabled'] = True
            request.session['auto_report_time'] = time
            
            return JsonResponse({
                'success': True,
                'message': f'Envoi automatique programmé pour {time} chaque jour'
            })
        else:
            # Désactiver la tâche automatique
            request.session['auto_report_enabled'] = False
            
            return JsonResponse({
                'success': True,
                'message': 'Envoi automatique désactivé'
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

def generate_daily_report_pdf(date_str):
    """
    Générer un rapport journalier en PDF avec design professionnel
    """
    from io import BytesIO
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from datetime import datetime, timedelta
    import os
    
    # Créer un buffer en mémoire
    buffer = BytesIO()
    
    # Créer le document PDF avec marges optimisées
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4,
        rightMargin=50,
        leftMargin=50,
        topMargin=50,
        bottomMargin=50
    )
    story = []
    styles = getSampleStyleSheet()
    
    # Styles personnalisés
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=15,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1f4e79'),
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=10,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#2c5282'),
        fontName='Helvetica-Bold'
    )
    
    section_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading3'],
        fontSize=12,
        spaceAfter=8,
        spaceBefore=12,
        textColor=colors.HexColor('#2d3748'),
        fontName='Helvetica-Bold',
        borderWidth=1,
        borderColor=colors.HexColor('#e2e8f0'),
        borderPadding=5,
        backColor=colors.HexColor('#f7fafc')
    )
    
    # En-tête avec logo et titre
    title = Paragraph("📊 RAPPORT JOURNALIER", title_style)
    story.append(title)
    
    subtitle = Paragraph("TS AIR CARGO - AGENCE MALI", subtitle_style)
    story.append(subtitle)
    
    # Ligne de séparation
    story.append(Spacer(1, 10))
    
    # Informations de base
    date_obj = datetime.strptime(date_str, '%Y-%m-%d')
    date_filter = date_obj.date()
    
    info_data = [
        ['📅 Date du rapport:', date_obj.strftime('%d/%m/%Y')],
        ['🕐 Généré le:', datetime.now().strftime('%d/%m/%Y à %H:%M')],
        ['🏢 Agence:', 'Mali - Bamako']
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 3*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#edf2f7')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2d3748')),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    
    story.append(info_table)
    story.append(Spacer(1, 12))
    
    # Statistiques du jour
    # Lots reçus
    lots_recus = ReceptionLot.objects.filter(date_reception__date=date_filter).count()
    
    # Colis livrés
    colis_livres = Livraison.objects.filter(
        date_livraison_effective__date=date_filter,
        statut='livree'
    ).count()
    
    # Dépenses
    depenses_total = Depense.objects.filter(
        date_depense=date_filter
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    # Revenus
    revenus_total = Livraison.objects.filter(
        date_livraison_effective__date=date_filter,
        statut='livree',
        montant_collecte__isnull=False
    ).aggregate(total=Sum('montant_collecte'))['total'] or 0
    
    # Bénéfice
    benefice_net = revenus_total - depenses_total
    
    # Vérifier s'il y a des données
    has_data = lots_recus > 0 or colis_livres > 0 or depenses_total > 0 or revenus_total > 0
    
    if has_data:
        # Section des indicateurs clés
        section_title = Paragraph("📈 INDICATEURS CLÉS DE PERFORMANCE", section_style)
        story.append(section_title)
        
        # Tableau des KPIs avec couleurs
        kpi_data = [
            ['📊 INDICATEUR', '📋 VALEUR', '📈 STATUT'],
            ['📦 Lots reçus', str(lots_recus), '✅ Traité' if lots_recus > 0 else '⚪ Aucun'],
            ['🚚 Colis livrés', str(colis_livres), '✅ Livré' if colis_livres > 0 else '⚪ Aucun'],
            ['💸 Dépenses totales', f"{depenses_total:,.0f} CFA", '🔴 Sortie' if depenses_total > 0 else '⚪ Aucune'],
            ['💰 Revenus de livraison', f"{revenus_total:,.0f} CFA", '🟢 Entrée' if revenus_total > 0 else '⚪ Aucun'],
            ['📊 Bénéfice net', f"{benefice_net:,.0f} CFA", 
             '🟢 Positif' if benefice_net > 0 else '🔴 Négatif' if benefice_net < 0 else '⚪ Équilibré']
        ]
        
        kpi_table = Table(kpi_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
        kpi_table.setStyle(TableStyle([
            # En-tête
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2b6cb0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            
            # Corps du tableau
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#2d3748')),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),
            ('ALIGN', (2, 1), (2, -1), 'CENTER'),
            
            # Alternance de couleurs
            ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#f7fafc')),
            ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#f7fafc')),
            ('BACKGROUND', (0, 6), (-1, 6), colors.HexColor('#f7fafc')),
            
            # Mise en forme spéciale pour le bénéfice
            ('BACKGROUND', (0, 5), (-1, 5), 
             colors.HexColor('#c6f6d5') if benefice_net > 0 else 
             colors.HexColor('#fed7d7') if benefice_net < 0 else colors.HexColor('#e2e8f0')),
            
            # Bordures et espacement
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e0')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(kpi_table)
        story.append(Spacer(1, 10))
        
        # Section détails des dépenses si il y en a
        if depenses_total > 0:
            section_depenses = Paragraph("💸 DÉTAIL DES DÉPENSES", section_style)
            story.append(section_depenses)
            
            depenses_detail = Depense.objects.filter(
                date_depense=date_filter
            ).values('type_depense', 'libelle', 'montant')[:10]  # Limiter à 10 entrées
            
            if depenses_detail:
                depenses_data = [['🏷️ TYPE', '📝 LIBELLÉ', '💰 MONTANT']]
                for dep in depenses_detail:
                    depenses_data.append([
                        dep['type_depense'].replace('_', ' ').title(),
                        dep['libelle'][:30] + '...' if len(dep['libelle']) > 30 else dep['libelle'],
                        f"{dep['montant']:,.0f} CFA"
                    ])
                
                depenses_table = Table(depenses_data, colWidths=[1.5*inch, 2.5*inch, 1.5*inch])
                depenses_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e53e3e')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#2d3748')),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('ALIGN', (0, 1), (1, -1), 'LEFT'),
                    ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
                    
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e0')),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]))
                
                story.append(depenses_table)
                story.append(Spacer(1, 8))
        
    else:
        # Aucune donnée disponible
        no_data_style = ParagraphStyle(
            'NoData',
            parent=styles['Normal'],
            fontSize=14,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#718096'),
            fontName='Helvetica'
        )
        
        no_data_msg = Paragraph(
            "📭 AUCUNE ACTIVITÉ ENREGISTRÉE<br/><br/>" + 
            "Aucune donnée n'a été trouvée pour cette date.<br/>" +
            "• Aucun lot reçu<br/>" +
            "• Aucun colis livré<br/>" +
            "• Aucune dépense enregistrée<br/>" +
            "• Aucun revenu généré<br/><br/>" +
            "Vérifiez que les données ont été correctement saisies.",
            no_data_style
        )
        story.append(no_data_msg)
        story.append(Spacer(1, 10))
    
    # Pied de page
    story.append(Spacer(1, 15))
    
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=9,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#718096'),
        fontName='Helvetica-Oblique'
    )
    
    footer = Paragraph(
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━<br/>" +
        f"📧 contact@ts-aircargo.com | 📞 +223 XX XX XX XX | 🌐 www.ts-aircargo.com<br/>" +
        f"© {datetime.now().year} TS Air Cargo Mali - Tous droits réservés",
        footer_style
    )
    story.append(footer)
    
    # Construire le PDF
    doc.build(story)
    
    # Récupérer le contenu
    pdf_content = buffer.getvalue()
    buffer.close()
    
    return pdf_content

def generate_monthly_report_pdf(month_str):
    """
    Générer un rapport mensuel en PDF avec design professionnel
    Format month_str: 'YYYY-MM'
    """
    from io import BytesIO
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from datetime import datetime, timedelta
    from calendar import monthrange
    import locale
    
    # Définir la locale en français si possible
    try:
        locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')
    except:
        pass
    
    # Créer un buffer en mémoire
    buffer = BytesIO()
    
    # Créer le document PDF avec marges
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4,
        rightMargin=72,
        leftMargin=72,
        topMargin=72,
        bottomMargin=18
    )
    story = []
    styles = getSampleStyleSheet()
    
    # Styles personnalisés
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1f4e79'),
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=20,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#2c5282'),
        fontName='Helvetica-Bold'
    )
    
    section_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading3'],
        fontSize=14,
        spaceAfter=15,
        spaceBefore=25,
        textColor=colors.HexColor('#2d3748'),
        fontName='Helvetica-Bold',
        borderWidth=1,
        borderColor=colors.HexColor('#e2e8f0'),
        borderPadding=8,
        backColor=colors.HexColor('#f7fafc')
    )
    
    # Traitement de la période
    year, month = map(int, month_str.split('-'))
    month_obj = datetime(year, month, 1)
    
    # Calculer les dates de début et fin du mois
    start_date = datetime(year, month, 1).date()
    _, last_day = monthrange(year, month)
    end_date = datetime(year, month, last_day).date()
    
    # En-tête avec logo et titre
    title = Paragraph("📊 RAPPORT MENSUEL", title_style)
    story.append(title)
    
    subtitle = Paragraph("TS AIR CARGO - AGENCE MALI", subtitle_style)
    story.append(subtitle)
    
    # Ligne de séparation
    story.append(Spacer(1, 20))
    
    # Informations de base
    month_name = month_obj.strftime('%B %Y').title()
    info_data = [
        ['📅 Période du rapport:', month_name],
        ['🕐 Généré le:', datetime.now().strftime('%d/%m/%Y à %H:%M')],
        ['🏢 Agence:', 'Mali - Bamako'],
        ['📈 Du:', start_date.strftime('%d/%m/%Y')],
        ['📈 Au:', end_date.strftime('%d/%m/%Y')]
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 3*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#edf2f7')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2d3748')),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    
    story.append(info_table)
    story.append(Spacer(1, 30))
    
    # Statistiques du mois
    # Lots reçus
    lots_recus = ReceptionLot.objects.filter(
        date_reception__date__gte=start_date,
        date_reception__date__lte=end_date
    ).count()
    
    # Colis livrés
    colis_livres = Livraison.objects.filter(
        date_livraison_effective__date__gte=start_date,
        date_livraison_effective__date__lte=end_date,
        statut='livree'
    ).count()
    
    # Dépenses
    depenses_total = Depense.objects.filter(
        date_depense__gte=start_date,
        date_depense__lte=end_date
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    # Revenus
    revenus_total = Livraison.objects.filter(
        date_livraison_effective__date__gte=start_date,
        date_livraison_effective__date__lte=end_date,
        statut='livree',
        montant_collecte__isnull=False
    ).aggregate(total=Sum('montant_collecte'))['total'] or 0
    
    # Bénéfice
    benefice_net = revenus_total - depenses_total
    
    # Moyennes journalières
    nombre_jours = (end_date - start_date).days + 1
    moy_lots_jour = lots_recus / nombre_jours if nombre_jours > 0 else 0
    moy_colis_jour = colis_livres / nombre_jours if nombre_jours > 0 else 0
    moy_revenus_jour = revenus_total / nombre_jours if nombre_jours > 0 else 0
    
    # Vérifier s'il y a des données
    has_data = lots_recus > 0 or colis_livres > 0 or depenses_total > 0 or revenus_total > 0
    
    if has_data:
        # Section des indicateurs clés
        section_title = Paragraph("📈 INDICATEURS CLÉS MENSUELS", section_style)
        story.append(section_title)
        
        # Tableau des KPIs avec couleurs
        kpi_data = [
            ['📊 INDICATEUR', '📋 VALEUR', '📊 MOYENNE/JOUR', '📈 STATUT'],
            ['📦 Lots reçus', str(lots_recus), f"{moy_lots_jour:.1f}", '✅ Actif' if lots_recus > 0 else '⚪ Inactif'],
            ['🚚 Colis livrés', str(colis_livres), f"{moy_colis_jour:.1f}", '✅ Performant' if colis_livres > 0 else '⚪ Aucun'],
            ['💸 Dépenses totales', f"{depenses_total:,.0f} CFA", f"{depenses_total/nombre_jours:,.0f} CFA", '🔴 Sortie' if depenses_total > 0 else '⚪ Aucune'],
            ['💰 Revenus totaux', f"{revenus_total:,.0f} CFA", f"{moy_revenus_jour:,.0f} CFA", '🟢 Entrée' if revenus_total > 0 else '⚪ Aucun'],
            ['📊 Bénéfice net', f"{benefice_net:,.0f} CFA", f"{benefice_net/nombre_jours:,.0f} CFA", 
             '🟢 Positif' if benefice_net > 0 else '🔴 Négatif' if benefice_net < 0 else '⚪ Équilibré']
        ]
        
        kpi_table = Table(kpi_data, colWidths=[2*inch, 1.3*inch, 1.3*inch, 1.4*inch])
        kpi_table.setStyle(TableStyle([
            # En-tête
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2b6cb0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            
            # Corps du tableau
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#2d3748')),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            
            # Alternance de couleurs
            ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#f7fafc')),
            ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#f7fafc')),
            ('BACKGROUND', (0, 6), (-1, 6), colors.HexColor('#f7fafc')),
            
            # Mise en forme spéciale pour le bénéfice
            ('BACKGROUND', (0, 5), (-1, 5), 
             colors.HexColor('#c6f6d5') if benefice_net > 0 else 
             colors.HexColor('#fed7d7') if benefice_net < 0 else colors.HexColor('#e2e8f0')),
            
            # Bordures et espacement
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e0')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        
        story.append(kpi_table)
        story.append(Spacer(1, 30))
        
        # Analyse des dépenses par type
        if depenses_total > 0:
            section_depenses = Paragraph("💸 ANALYSE DES DÉPENSES PAR CATÉGORIE", section_style)
            story.append(section_depenses)
            
            depenses_par_type = Depense.objects.filter(
                date_depense__gte=start_date,
                date_depense__lte=end_date
            ).values('type_depense').annotate(
                total=Sum('montant'),
                count=Count('id')
            ).order_by('-total')
            
            if depenses_par_type:
                depenses_data = [['🏷️ CATÉGORIE', '💰 MONTANT TOTAL', '📊 NOMBRE', '📈 % DU TOTAL']]
                for dep in depenses_par_type:
                    pourcentage = (dep['total'] / depenses_total * 100) if depenses_total > 0 else 0
                    depenses_data.append([
                        dep['type_depense'].replace('_', ' ').title(),
                        f"{dep['total']:,.0f} CFA",
                        str(dep['count']),
                        f"{pourcentage:.1f}%"
                    ])
                
                depenses_table = Table(depenses_data, colWidths=[2*inch, 1.5*inch, 1*inch, 1*inch])
                depenses_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e53e3e')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#2d3748')),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                    ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
                    
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e0')),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 8),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ]))
                
                story.append(depenses_table)
                story.append(Spacer(1, 20))
    
    else:
        # Aucune donnée disponible
        no_data_style = ParagraphStyle(
            'NoData',
            parent=styles['Normal'],
            fontSize=14,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#718096'),
            fontName='Helvetica'
        )
        
        no_data_msg = Paragraph(
            f"📭 AUCUNE ACTIVITÉ ENREGISTRÉE POUR {month_name.upper()}<br/><br/>" + 
            "Aucune donnée n'a été trouvée pour cette période.<br/>" +
            "• Aucun lot reçu<br/>" +
            "• Aucun colis livré<br/>" +
            "• Aucune dépense enregistrée<br/>" +
            "• Aucun revenu généré<br/><br/>" +
            "Vérifiez que les données ont été correctement saisies.",
            no_data_style
        )
        story.append(no_data_msg)
        story.append(Spacer(1, 30))
    
    # Pied de page
    story.append(Spacer(1, 50))
    
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#718096'),
        fontName='Helvetica-Oblique'
    )
    
    footer = Paragraph(
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━<br/>" +
        f"📧 contact@ts-aircargo.com | 📞 +223 XX XX XX XX | 🌐 www.ts-aircargo.com<br/>" +
        f"© {datetime.now().year} TS Air Cargo Mali - Tous droits réservés",
        footer_style
    )
    story.append(footer)
    
    # Construire le PDF
    doc.build(story)
    
    # Récupérer le contenu
    pdf_content = buffer.getvalue()
    buffer.close()
    
    return pdf_content

def generate_yearly_report_pdf(year_str):
    """
    Générer un rapport annuel en PDF avec design professionnel
    Format year_str: 'YYYY'
    """
    from io import BytesIO
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from datetime import datetime, timedelta
    
    # Créer un buffer en mémoire
    buffer = BytesIO()
    
    # Créer le document PDF avec marges
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4,
        rightMargin=72,
        leftMargin=72,
        topMargin=72,
        bottomMargin=18
    )
    story = []
    styles = getSampleStyleSheet()
    
    # Styles personnalisés
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1f4e79'),
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=20,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#2c5282'),
        fontName='Helvetica-Bold'
    )
    
    section_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading3'],
        fontSize=14,
        spaceAfter=15,
        spaceBefore=25,
        textColor=colors.HexColor('#2d3748'),
        fontName='Helvetica-Bold',
        borderWidth=1,
        borderColor=colors.HexColor('#e2e8f0'),
        borderPadding=8,
        backColor=colors.HexColor('#f7fafc')
    )
    
    # Traitement de l'année
    year = int(year_str)
    start_date = datetime(year, 1, 1).date()
    end_date = datetime(year, 12, 31).date()
    
    # En-tête avec logo et titre
    title = Paragraph("📊 RAPPORT ANNUEL", title_style)
    story.append(title)
    
    subtitle = Paragraph("TS AIR CARGO - AGENCE MALI", subtitle_style)
    story.append(subtitle)
    
    # Ligne de séparation
    story.append(Spacer(1, 20))
    
    # Informations de base
    info_data = [
        ['📅 Année du rapport:', str(year)],
        ['🕐 Généré le:', datetime.now().strftime('%d/%m/%Y à %H:%M')],
        ['🏢 Agence:', 'Mali - Bamako'],
        ['📈 Période:', f"Du 01/01/{year} au 31/12/{year}"]
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 3*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#edf2f7')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2d3748')),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    
    story.append(info_table)
    story.append(Spacer(1, 30))
    
    # Statistiques de l'année
    # Lots reçus
    lots_recus = ReceptionLot.objects.filter(
        date_reception__date__gte=start_date,
        date_reception__date__lte=end_date
    ).count()
    
    # Colis livrés
    colis_livres = Livraison.objects.filter(
        date_livraison_effective__date__gte=start_date,
        date_livraison_effective__date__lte=end_date,
        statut='livree'
    ).count()
    
    # Dépenses
    depenses_total = Depense.objects.filter(
        date_depense__gte=start_date,
        date_depense__lte=end_date
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    # Revenus
    revenus_total = Livraison.objects.filter(
        date_livraison_effective__date__gte=start_date,
        date_livraison_effective__date__lte=end_date,
        statut='livree',
        montant_collecte__isnull=False
    ).aggregate(total=Sum('montant_collecte'))['total'] or 0
    
    # Bénéfice
    benefice_net = revenus_total - depenses_total
    
    # Moyennes mensuelles
    moy_lots_mois = lots_recus / 12
    moy_colis_mois = colis_livres / 12
    moy_revenus_mois = revenus_total / 12
    moy_depenses_mois = depenses_total / 12
    
    # Vérifier s'il y a des données
    has_data = lots_recus > 0 or colis_livres > 0 or depenses_total > 0 or revenus_total > 0
    
    if has_data:
        # Section des indicateurs clés
        section_title = Paragraph("📈 BILAN ANNUEL - INDICATEURS CLÉS", section_style)
        story.append(section_title)
        
        # Tableau des KPIs avec couleurs
        kpi_data = [
            ['📊 INDICATEUR', '📋 TOTAL ANNUEL', '📊 MOYENNE/MOIS', '📈 PERFORMANCE'],
            ['📦 Lots reçus', str(lots_recus), f"{moy_lots_mois:.1f}", '✅ Excellent' if lots_recus > 100 else '🟡 Moyen' if lots_recus > 50 else '🔴 Faible'],
            ['🚚 Colis livrés', str(colis_livres), f"{moy_colis_mois:.1f}", '✅ Excellent' if colis_livres > 500 else '🟡 Moyen' if colis_livres > 200 else '🔴 Faible'],
            ['💸 Dépenses totales', f"{depenses_total:,.0f} CFA", f"{moy_depenses_mois:,.0f} CFA", '🔴 Élevé' if depenses_total > 1000000 else '🟡 Modéré'],
            ['💰 Revenus totaux', f"{revenus_total:,.0f} CFA", f"{moy_revenus_mois:,.0f} CFA", '✅ Excellent' if revenus_total > 2000000 else '🟡 Moyen'],
            ['📊 Bénéfice net', f"{benefice_net:,.0f} CFA", f"{benefice_net/12:,.0f} CFA", 
             '🟢 Très Positif' if benefice_net > 1000000 else '✅ Positif' if benefice_net > 0 else '🔴 Négatif']
        ]
        
        kpi_table = Table(kpi_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        kpi_table.setStyle(TableStyle([
            # En-tête
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2b6cb0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            
            # Corps du tableau
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#2d3748')),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            
            # Alternance de couleurs
            ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#f7fafc')),
            ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#f7fafc')),
            ('BACKGROUND', (0, 6), (-1, 6), colors.HexColor('#f7fafc')),
            
            # Mise en forme spéciale pour le bénéfice
            ('BACKGROUND', (0, 5), (-1, 5), 
             colors.HexColor('#c6f6d5') if benefice_net > 0 else 
             colors.HexColor('#fed7d7') if benefice_net < 0 else colors.HexColor('#e2e8f0')),
            
            # Bordures et espacement
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e0')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        
        story.append(kpi_table)
        story.append(Spacer(1, 30))
        
        # Évolution mensuelle (tableau récapitulatif)
        section_evolution = Paragraph("📈 ÉVOLUTION MENSUELLE", section_style)
        story.append(section_evolution)
        
        # Créer un tableau avec l'évolution mois par mois
        evolution_data = [['MOIS', 'LOTS REÇUS', 'COLIS LIVRÉS', 'REVENUS (CFA)', 'DÉPENSES (CFA)']]
        
        for month in range(1, 13):
            month_start = datetime(year, month, 1).date()
            if month == 12:
                month_end = datetime(year, 12, 31).date()
            else:
                month_end = datetime(year, month + 1, 1).date() - timedelta(days=1)
            
            # Statistiques du mois
            lots_mois = ReceptionLot.objects.filter(
                date_reception__date__gte=month_start,
                date_reception__date__lte=month_end
            ).count()
            
            colis_mois = Livraison.objects.filter(
                date_livraison_effective__date__gte=month_start,
                date_livraison_effective__date__lte=month_end,
                statut='livree'
            ).count()
            
            revenus_mois = Livraison.objects.filter(
                date_livraison_effective__date__gte=month_start,
                date_livraison_effective__date__lte=month_end,
                statut='livree',
                montant_collecte__isnull=False
            ).aggregate(total=Sum('montant_collecte'))['total'] or 0
            
            depenses_mois = Depense.objects.filter(
                date_depense__gte=month_start,
                date_depense__lte=month_end
            ).aggregate(total=Sum('montant'))['total'] or 0
            
            month_name = datetime(year, month, 1).strftime('%B')[:3].title()
            evolution_data.append([
                month_name,
                str(lots_mois),
                str(colis_mois),
                f"{revenus_mois:,.0f}" if revenus_mois > 0 else "-",
                f"{depenses_mois:,.0f}" if depenses_mois > 0 else "-"
            ])
        
        evolution_table = Table(evolution_data, colWidths=[1*inch, 1*inch, 1*inch, 1.5*inch, 1.5*inch])
        evolution_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#38a169')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#2d3748')),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e0')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        
        story.append(evolution_table)
        story.append(Spacer(1, 20))
    
    else:
        # Aucune donnée disponible
        no_data_style = ParagraphStyle(
            'NoData',
            parent=styles['Normal'],
            fontSize=14,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#718096'),
            fontName='Helvetica'
        )
        
        no_data_msg = Paragraph(
            f"📭 AUCUNE ACTIVITÉ ENREGISTRÉE POUR L'ANNÉE {year}<br/><br/>" + 
            "Aucune donnée n'a été trouvée pour cette année.<br/>" +
            "• Aucun lot reçu<br/>" +
            "• Aucun colis livré<br/>" +
            "• Aucune dépense enregistrée<br/>" +
            "• Aucun revenu généré<br/><br/>" +
            "Vérifiez que les données ont été correctement saisies.",
            no_data_style
        )
        story.append(no_data_msg)
        story.append(Spacer(1, 30))
    
    # Pied de page
    story.append(Spacer(1, 50))
    
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#718096'),
        fontName='Helvetica-Oblique'
    )
    
    footer = Paragraph(
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━<br/>" +
        f"📧 contact@ts-aircargo.com | 📞 +223 XX XX XX XX | 🌐 www.ts-aircargo.com<br/>" +
        f"© {datetime.now().year} TS Air Cargo Mali - Tous droits réservés",
        footer_style
    )
    story.append(footer)
    
    # Construire le PDF
    doc.build(story)
    
    # Récupérer le contenu
    pdf_content = buffer.getvalue()
    buffer.close()
    
    return pdf_content


@agent_mali_required
def export_depenses_excel(request):
    """
    Exporter les dépenses en format Excel avec formatage professionnel
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from datetime import datetime
    
    # Récupération des filtres
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    type_depense = request.GET.get('type_depense')
    
    # Dates par défaut
    today = timezone.now().date()
    if not date_debut:
        date_debut = today - timedelta(days=30)
    else:
        try:
            date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
        except ValueError:
            date_debut = today - timedelta(days=30)
    
    if not date_fin:
        date_fin = today
    else:
        try:
            date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError:
            date_fin = today
    
    # Récupération des dépenses
    depenses_query = Depense.objects.filter(
        date_depense__gte=date_debut,
        date_depense__lte=date_fin
    )
    
    if type_depense:
        depenses_query = depenses_query.filter(type_depense=type_depense)
    
    depenses = depenses_query.select_related('cree_par').order_by('-date_depense')
    
    # Création du workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapport Dépenses"
    
    # Styles pour le formatage
    header_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF6B35', end_color='FF6B35', fill_type='solid')
    
    subheader_font = Font(name='Arial', size=12, bold=True, color='2D3748')
    subheader_fill = PatternFill(start_color='F7FAFC', end_color='F7FAFC', fill_type='solid')
    
    data_font = Font(name='Arial', size=10)
    number_font = Font(name='Arial', size=10, bold=True)
    
    border = Border(
        left=Side(border_style='thin', color='E2E8F0'),
        right=Side(border_style='thin', color='E2E8F0'),
        top=Side(border_style='thin', color='E2E8F0'),
        bottom=Side(border_style='thin', color='E2E8F0')
    )
    
    # En-tête principal
    ws['A1'] = 'RAPPORT DES DÉPENSES - TS AIR CARGO MALI'
    ws.merge_cells('A1:G1')
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Sous-titre avec période
    ws['A2'] = f'Période: {date_debut.strftime("%d/%m/%Y")} au {date_fin.strftime("%d/%m/%Y")}'
    ws.merge_cells('A2:G2')
    ws['A2'].font = subheader_font
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Date de génération
    ws['A3'] = f'Généré le: {datetime.now().strftime("%d/%m/%Y à %H:%M")}'
    ws.merge_cells('A3:G3')
    ws['A3'].font = data_font
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Résumé
    total_depenses = depenses.aggregate(total=Sum('montant'))['total'] or 0
    ws['A5'] = f'Total des dépenses: {total_depenses:,.0f} FCFA'
    ws['A5'].font = subheader_font
    ws['A6'] = f'Nombre de dépenses: {depenses.count()}'
    ws['A6'].font = subheader_font
    
    # En-têtes des colonnes
    headers = ['Date', 'Type', 'Description', 'Montant (FCFA)', 'Créé par', 'Date création', 'Statut']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=8, column=col_idx, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Données des dépenses
    for row_idx, depense in enumerate(depenses, start=9):
        row_data = [
            depense.date_depense.strftime('%d/%m/%Y'),
            depense.get_type_depense_display(),
            depense.description,
            float(depense.montant),
            f"{depense.cree_par.first_name} {depense.cree_par.last_name}" if depense.cree_par else "N/A",
            depense.date_creation.strftime('%d/%m/%Y %H:%M'),
            'Validé'
        ]
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col_idx == 4:  # Colonne montant
                cell.font = number_font
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.font = data_font
    
    # Ajuster la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Préparation de la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="depenses_{date_debut}_{date_fin}.xlsx"'
    
    wb.save(response)
    return response


@agent_mali_required
def export_rapport_cargo_excel(request):
    """
    Exporter les rapports cargo en format Excel
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    from datetime import datetime
    
    # Récupération des filtres
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    # Dates par défaut
    today = timezone.now().date()
    if not date_debut:
        date_debut = today - timedelta(days=30)
    else:
        try:
            date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
        except ValueError:
            date_debut = today - timedelta(days=30)
    
    if not date_fin:
        date_fin = today
    else:
        try:
            date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError:
            date_fin = today
    
    # Récupération des colis cargo
    colis_cargo = Colis.objects.filter(
        type_transport='cargo',
        date_creation__gte=date_debut,
        date_creation__lte=date_fin
    ).select_related('client__user', 'lot').order_by('-date_creation')
    
    # Récupération des lots avec colis cargo
    lots_cargo = Lot.objects.filter(
        colis__type_transport='cargo',
        date_creation__gte=date_debut,
        date_creation__lte=date_fin
    ).distinct().prefetch_related('colis_set').order_by('-date_creation')
    
    # Création du workbook Excel
    wb = openpyxl.Workbook()
    
    # Feuille 1: Résumé Cargo
    ws1 = wb.active
    ws1.title = "Résumé Cargo"
    
    # Styles
    header_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF6B35', end_color='FF6B35', fill_type='solid')
    subheader_font = Font(name='Arial', size=12, bold=True, color='2D3748')
    subheader_fill = PatternFill(start_color='F7FAFC', end_color='F7FAFC', fill_type='solid')
    data_font = Font(name='Arial', size=10)
    number_font = Font(name='Arial', size=10, bold=True)
    border = Border(
        left=Side(border_style='thin', color='E2E8F0'),
        right=Side(border_style='thin', color='E2E8F0'),
        top=Side(border_style='thin', color='E2E8F0'),
        bottom=Side(border_style='thin', color='E2E8F0')
    )
    
    # En-tête
    ws1['A1'] = 'RAPPORT CARGO - TS AIR CARGO'
    ws1.merge_cells('A1:F1')
    ws1['A1'].font = header_font
    ws1['A1'].fill = header_fill
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Statistiques générales
    total_colis = colis_cargo.count()
    valeur_totale = colis_cargo.aggregate(total=Sum('prix_calcule'))['total'] or 0
    poids_total = colis_cargo.aggregate(total=Sum('poids'))['total'] or 0
    
    stats_data = [
        ['Indicateur', 'Valeur'],
        ['Total Colis Cargo', total_colis],
        ['Valeur Totale (FCFA)', f"{valeur_totale:,.0f}"],
        ['Poids Total (Kg)', f"{poids_total:,.1f}"],
        ['Nombre de Lots', lots_cargo.count()],
    ]
    
    for row_idx, row_data in enumerate(stats_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if row_idx == 3:
                cell.font = subheader_font
                cell.fill = subheader_fill
            else:
                cell.font = data_font
    
    # Feuille 2: Détail des Colis Cargo
    ws2 = wb.create_sheet(title="Colis Cargo")
    
    # En-têtes des colis
    colis_headers = ['Code Suivi', 'Client', 'Description', 'Poids (Kg)', 'Valeur (FCFA)', 'Statut', 'Date Création', 'Lot']
    for col_idx, header in enumerate(colis_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = border
    
    # Données des colis
    for row_idx, colis in enumerate(colis_cargo, start=2):
        row_data = [
            colis.code_suivi,
            f"{colis.client.user.first_name} {colis.client.user.last_name}" if colis.client and colis.client.user else "N/A",
            colis.description_contenu,
            float(colis.poids) if colis.poids else 0,
            float(colis.prix_calcule) if colis.prix_calcule else 0,
            colis.get_statut_display(),
            colis.date_creation.strftime('%d/%m/%Y'),
            colis.lot.numero_lot if colis.lot else "Aucun"
        ]
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col_idx in [4, 5]:  # Colonnes numériques
                cell.font = number_font
                cell.number_format = '#,##0.00' if col_idx == 4 else '#,##0'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.font = data_font
    
    # Feuille 3: Lots Cargo
    ws3 = wb.create_sheet(title="Lots Cargo")
    
    # En-têtes des lots
    lots_headers = ['Numéro Lot', 'Nb Colis', 'Poids Total (Kg)', 'Prix Transport (FCFA)', 'Statut', 'Date Création', 'Date Expédition']
    for col_idx, header in enumerate(lots_headers, start=1):
        cell = ws3.cell(row=1, column=col_idx, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = border
    
    # Données des lots
    for row_idx, lot in enumerate(lots_cargo, start=2):
        nb_colis_cargo = lot.colis_set.filter(type_transport='cargo').count()
        poids_lot = lot.colis_set.filter(type_transport='cargo').aggregate(total=Sum('poids'))['total'] or 0
        
        row_data = [
            lot.numero_lot,
            nb_colis_cargo,
            float(poids_lot),
            float(lot.prix_transport) if lot.prix_transport else 0,
            lot.get_statut_display(),
            lot.date_creation.strftime('%d/%m/%Y'),
            lot.date_expedition.strftime('%d/%m/%Y') if lot.date_expedition else "Non expédié"
        ]
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col_idx in [2, 3, 4]:  # Colonnes numériques
                cell.font = number_font
                if col_idx == 3:
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.font = data_font
    
    # Ajuster les largeurs des colonnes pour toutes les feuilles
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    # Préparation de la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="rapport_cargo_{date_debut}_{date_fin}.xlsx"'
    
    wb.save(response)
    return response


@agent_mali_required
def export_rapport_express_excel(request):
    """
    Exporter les rapports express en format Excel
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    from datetime import datetime
    
    # Récupération des filtres
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    # Dates par défaut
    today = timezone.now().date()
    if not date_debut:
        date_debut = today - timedelta(days=30)
    else:
        try:
            date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
        except ValueError:
            date_debut = today - timedelta(days=30)
    
    if not date_fin:
        date_fin = today
    else:
        try:
            date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError:
            date_fin = today
    
    # Récupération des colis express
    colis_express = Colis.objects.filter(
        type_transport='express',
        date_creation__gte=date_debut,
        date_creation__lte=date_fin
    ).select_related('client__user', 'lot').order_by('-date_creation')
    
    # Récupération des lots avec colis express
    lots_express = Lot.objects.filter(
        colis__type_transport='express',
        date_creation__gte=date_debut,
        date_creation__lte=date_fin
    ).distinct().prefetch_related('colis_set').order_by('-date_creation')
    
    # Création du workbook Excel
    wb = openpyxl.Workbook()
    
    # Feuille 1: Résumé Express
    ws1 = wb.active
    ws1.title = "Résumé Express"
    
    # Styles
    header_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')  # Vert pour express
    subheader_font = Font(name='Arial', size=12, bold=True, color='2D3748')
    subheader_fill = PatternFill(start_color='F7FAFC', end_color='F7FAFC', fill_type='solid')
    data_font = Font(name='Arial', size=10)
    number_font = Font(name='Arial', size=10, bold=True)
    border = Border(
        left=Side(border_style='thin', color='E2E8F0'),
        right=Side(border_style='thin', color='E2E8F0'),
        top=Side(border_style='thin', color='E2E8F0'),
        bottom=Side(border_style='thin', color='E2E8F0')
    )
    
    # En-tête
    ws1['A1'] = 'RAPPORT EXPRESS - TS AIR CARGO'
    ws1.merge_cells('A1:F1')
    ws1['A1'].font = header_font
    ws1['A1'].fill = header_fill
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Statistiques générales
    total_colis = colis_express.count()
    valeur_totale = colis_express.aggregate(total=Sum('prix_calcule'))['total'] or 0
    poids_total = colis_express.aggregate(total=Sum('poids'))['total'] or 0
    
    # Calcul du délai moyen de livraison
    colis_livres = colis_express.filter(statut='livre')
    delai_moyen = 0
    if colis_livres.exists():
        delais = []
        for colis in colis_livres:
            livraisons = colis.livraisons.filter(statut='livree')
            if livraisons.exists():
                livraison = livraisons.first()
                if livraison.date_livraison_effective:
                    delai = (livraison.date_livraison_effective.date() - colis.date_creation.date()).days
                    delais.append(delai)
        if delais:
            delai_moyen = sum(delais) / len(delais)
    
    stats_data = [
        ['Indicateur', 'Valeur'],
        ['Total Colis Express', total_colis],
        ['Valeur Totale (FCFA)', f"{valeur_totale:,.0f}"],
        ['Poids Total (Kg)', f"{poids_total:,.1f}"],
        ['Nombre de Lots', lots_express.count()],
        ['Délai Moyen Livraison (jours)', f"{delai_moyen:.1f}"],
        ['Taux de Livraison (%)', f"{(colis_livres.count() / total_colis * 100):.1f}" if total_colis > 0 else "0.0"],
    ]
    
    for row_idx, row_data in enumerate(stats_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if row_idx == 3:
                cell.font = subheader_font
                cell.fill = subheader_fill
            else:
                cell.font = data_font
    
    # Feuille 2: Détail des Colis Express
    ws2 = wb.create_sheet(title="Colis Express")
    
    # En-têtes des colis
    colis_headers = ['Code Suivi', 'Client', 'Description', 'Poids (Kg)', 'Valeur (FCFA)', 'Statut', 'Date Création', 'Date Livraison', 'Délai (jours)', 'Lot']
    for col_idx, header in enumerate(colis_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = border
    
    # Données des colis
    for row_idx, colis in enumerate(colis_express, start=2):
        # Calcul du délai
        delai = ""
        date_livraison = ""
        livraisons = colis.livraisons.filter(statut='livree')
        if livraisons.exists():
            livraison = livraisons.first()
            if livraison.date_livraison_effective:
                date_livraison = livraison.date_livraison_effective.strftime('%d/%m/%Y')
                delai = (livraison.date_livraison_effective.date() - colis.date_creation.date()).days
        
        row_data = [
            colis.code_suivi,
            f"{colis.client.user.first_name} {colis.client.user.last_name}" if colis.client and colis.client.user else "N/A",
            colis.description_contenu,
            float(colis.poids) if colis.poids else 0,
            float(colis.prix_calcule) if colis.prix_calcule else 0,
            colis.get_statut_display(),
            colis.date_creation.strftime('%d/%m/%Y'),
            date_livraison if date_livraison else "Non livré",
            delai if delai != "" else "N/A",
            colis.lot.numero_lot if colis.lot else "Aucun"
        ]
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col_idx in [4, 5, 9]:  # Colonnes numériques
                cell.font = number_font
                if col_idx == 4:
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.font = data_font
    
    # Feuille 3: Lots Express
    ws3 = wb.create_sheet(title="Lots Express")
    
    # En-têtes des lots
    lots_headers = ['Numéro Lot', 'Nb Colis', 'Poids Total (Kg)', 'Prix Transport (FCFA)', 'Frais Douane (FCFA)', 'Statut', 'Date Création', 'Date Expédition']
    for col_idx, header in enumerate(lots_headers, start=1):
        cell = ws3.cell(row=1, column=col_idx, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = border
    
    # Données des lots
    for row_idx, lot in enumerate(lots_express, start=2):
        nb_colis_express = lot.colis_set.filter(type_transport='express').count()
        poids_lot = lot.colis_set.filter(type_transport='express').aggregate(total=Sum('poids'))['total'] or 0
        
        row_data = [
            lot.numero_lot,
            nb_colis_express,
            float(poids_lot),
            float(lot.prix_transport) if lot.prix_transport else 0,
            float(lot.frais_douane) if hasattr(lot, 'frais_douane') and lot.frais_douane else 0,
            lot.get_statut_display(),
            lot.date_creation.strftime('%d/%m/%Y'),
            lot.date_expedition.strftime('%d/%m/%Y') if lot.date_expedition else "Non expédié"
        ]
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col_idx in [2, 3, 4, 5]:  # Colonnes numériques
                cell.font = number_font
                if col_idx == 3:
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.font = data_font
    
    # Ajuster les largeurs des colonnes pour toutes les feuilles
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    # Préparation de la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="rapport_express_{date_debut}_{date_fin}.xlsx"'
    
    wb.save(response)
    return response


@agent_mali_required
def export_rapport_bateau_excel(request):
    """
    Exporter les rapports bateau en format Excel
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    from datetime import datetime
    
    # Récupération des filtres
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    # Dates par défaut
    today = timezone.now().date()
    if not date_debut:
        date_debut = today - timedelta(days=30)
    else:
        try:
            date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
        except ValueError:
            date_debut = today - timedelta(days=30)
    
    if not date_fin:
        date_fin = today
    else:
        try:
            date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError:
            date_fin = today
    
    # Récupération des colis bateau
    colis_bateau = Colis.objects.filter(
        type_transport='bateau',
        date_creation__gte=date_debut,
        date_creation__lte=date_fin
    ).select_related('client__user', 'lot').order_by('-date_creation')
    
    # Récupération des lots avec colis bateau
    lots_bateau = Lot.objects.filter(
        colis__type_transport='bateau',
        date_creation__gte=date_debut,
        date_creation__lte=date_fin
    ).distinct().prefetch_related('colis_set').order_by('-date_creation')
    
    # Création du workbook Excel
    wb = openpyxl.Workbook()
    
    # Feuille 1: Résumé Bateau
    ws1 = wb.active
    ws1.title = "Résumé Bateau"
    
    # Styles
    header_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='007BFF', end_color='007BFF', fill_type='solid')  # Bleu pour bateau
    subheader_font = Font(name='Arial', size=12, bold=True, color='2D3748')
    subheader_fill = PatternFill(start_color='F7FAFC', end_color='F7FAFC', fill_type='solid')
    data_font = Font(name='Arial', size=10)
    number_font = Font(name='Arial', size=10, bold=True)
    border = Border(
        left=Side(border_style='thin', color='E2E8F0'),
        right=Side(border_style='thin', color='E2E8F0'),
        top=Side(border_style='thin', color='E2E8F0'),
        bottom=Side(border_style='thin', color='E2E8F0')
    )
    
    # En-tête
    ws1['A1'] = 'RAPPORT BATEAU - TS AIR CARGO'
    ws1.merge_cells('A1:F1')
    ws1['A1'].font = header_font
    ws1['A1'].fill = header_fill
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Statistiques générales
    total_colis = colis_bateau.count()
    valeur_totale = colis_bateau.aggregate(total=Sum('prix_calcule'))['total'] or 0
    poids_total = colis_bateau.aggregate(total=Sum('poids'))['total'] or 0
    
    # Calcul du délai moyen de livraison
    colis_livres = colis_bateau.filter(statut='livre')
    delai_moyen = 0
    if colis_livres.exists():
        delais = []
        for colis in colis_livres:
            livraisons = colis.livraisons.filter(statut='livree')
            if livraisons.exists():
                livraison = livraisons.first()
                if livraison.date_livraison_effective:
                    delai = (livraison.date_livraison_effective.date() - colis.date_creation.date()).days
                    delais.append(delai)
        if delais:
            delai_moyen = sum(delais) / len(delais)
    
    stats_data = [
        ['Indicateur', 'Valeur'],
        ['Total Colis Bateau', total_colis],
        ['Valeur Totale (FCFA)', f"{valeur_totale:,.0f}"],
        ['Poids Total (Kg)', f"{poids_total:,.1f}"],
        ['Nombre de Lots', lots_bateau.count()],
        ['Délai Moyen Livraison (jours)', f"{delai_moyen:.1f}"],
        ['Taux de Livraison (%)', f"{(colis_livres.count() / total_colis * 100):.1f}" if total_colis > 0 else "0.0"],
    ]
    
    for row_idx, row_data in enumerate(stats_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if row_idx == 3:
                cell.font = subheader_font
                cell.fill = subheader_fill
            else:
                cell.font = data_font
    
    # Feuille 2: Détail des Colis Bateau
    ws2 = wb.create_sheet(title="Colis Bateau")
    
    # En-têtes des colis
    colis_headers = ['Code Suivi', 'Client', 'Description', 'Poids (Kg)', 'Valeur (FCFA)', 'Statut', 'Date Création', 'Date Livraison', 'Délai (jours)', 'Lot']
    for col_idx, header in enumerate(colis_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = border
    
    # Données des colis
    for row_idx, colis in enumerate(colis_bateau, start=2):
        # Calcul du délai
        delai = ""
        date_livraison = ""
        livraisons = colis.livraisons.filter(statut='livree')
        if livraisons.exists():
            livraison = livraisons.first()
            if livraison.date_livraison_effective:
                date_livraison = livraison.date_livraison_effective.strftime('%d/%m/%Y')
                delai = (livraison.date_livraison_effective.date() - colis.date_creation.date()).days
        
        row_data = [
            colis.code_suivi,
            f"{colis.client.user.first_name} {colis.client.user.last_name}" if colis.client and colis.client.user else "N/A",
            colis.description_contenu,
            float(colis.poids) if colis.poids else 0,
            float(colis.prix_calcule) if colis.prix_calcule else 0,
            colis.get_statut_display(),
            colis.date_creation.strftime('%d/%m/%Y'),
            date_livraison if date_livraison else "Non livré",
            delai if delai != "" else "N/A",
            colis.lot.numero_lot if colis.lot else "Aucun"
        ]
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col_idx in [4, 5, 9]:  # Colonnes numériques
                cell.font = number_font
                if col_idx == 4:
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.font = data_font
    
    # Feuille 3: Lots Bateau
    ws3 = wb.create_sheet(title="Lots Bateau")
    
    # En-têtes des lots
    lots_headers = ['Numéro Lot', 'Nb Colis', 'Poids Total (Kg)', 'Prix Transport (FCFA)', 'Frais Douane (FCFA)', 'Statut', 'Date Création', 'Date Expédition']
    for col_idx, header in enumerate(lots_headers, start=1):
        cell = ws3.cell(row=1, column=col_idx, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = border
    
    # Données des lots
    for row_idx, lot in enumerate(lots_bateau, start=2):
        nb_colis_bateau = lot.colis_set.filter(type_transport='bateau').count()
        poids_lot = lot.colis_set.filter(type_transport='bateau').aggregate(total=Sum('poids'))['total'] or 0
        
        row_data = [
            lot.numero_lot,
            nb_colis_bateau,
            float(poids_lot),
            float(lot.prix_transport) if lot.prix_transport else 0,
            float(lot.frais_douane) if hasattr(lot, 'frais_douane') and lot.frais_douane else 0,
            lot.get_statut_display(),
            lot.date_creation.strftime('%d/%m/%Y'),
            lot.date_expedition.strftime('%d/%m/%Y') if lot.date_expedition else "Non expédié"
        ]
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col_idx in [2, 3, 4, 5]:  # Colonnes numériques
                cell.font = number_font
                if col_idx == 3:
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.font = data_font
    
    # Ajuster les largeurs des colonnes pour toutes les feuilles
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    # Préparation de la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="rapport_bateau_{date_debut}_{date_fin}.xlsx"'
    
    wb.save(response)
    return response
