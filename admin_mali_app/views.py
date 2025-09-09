from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal

from .models import TransfertArgent
from agent_chine_app.models import Lot, Colis
from agent_mali_app.models import Depense
from authentication.models import CustomUser


def admin_mali_required(view_func):
    """
    Décorateur pour vérifier que l'utilisateur est un admin Mali
    """
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('authentication:role_based_login', role='admin_mali')
        if not request.user.is_admin_mali:
            messages.error(request, "Accès refusé. Vous devez être administrateur Mali.")
            return redirect('authentication:home')
        return view_func(request, *args, **kwargs)
    return wrapper


@admin_mali_required
def dashboard(request):
    """
    Dashboard principal de l'admin Mali avec statistiques complètes
    """
    # Période pour les statistiques
    today = timezone.now().date()
    yesterday = today - timedelta(days=1)
    last_7_days = today - timedelta(days=7)
    last_30_days = today - timedelta(days=30)
    current_month = today.replace(day=1)
    last_month = (current_month - timedelta(days=1)).replace(day=1)
    
    # === STATISTIQUES DES TRANSFERTS D'ARGENT ===
    transferts_stats = {
        'total': TransfertArgent.objects.count(),
        'aujourd_hui': TransfertArgent.objects.filter(date_initiation__date=today).count(),
        'hier': TransfertArgent.objects.filter(date_initiation__date=yesterday).count(),
        '7_derniers_jours': TransfertArgent.objects.filter(date_initiation__date__gte=last_7_days).count(),
        'ce_mois': TransfertArgent.objects.filter(date_initiation__gte=current_month).count(),
        'mois_precedent': TransfertArgent.objects.filter(
            date_initiation__gte=last_month,
            date_initiation__lt=current_month
        ).count(),
        
        # Par statut
        'en_attente': TransfertArgent.objects.filter(statut='initie').count(),
        'envoyes': TransfertArgent.objects.filter(statut='envoye').count(),
        'confirmes': TransfertArgent.objects.filter(statut='confirme_chine').count(),
        'annules': TransfertArgent.objects.filter(statut='annule').count(),
    }
    
    # === MONTANTS FINANCIERS ===
    montants_transferts = TransfertArgent.objects.aggregate(
        # Montants totaux
        total_fcfa=Sum('montant_fcfa'),
        total_frais=Sum('frais_transfert'),
        
        # Montants journaliers
        aujourd_hui_fcfa=Sum('montant_fcfa', filter=Q(date_initiation__date=today)),
        aujourd_hui_frais=Sum('frais_transfert', filter=Q(date_initiation__date=today)),
        hier_fcfa=Sum('montant_fcfa', filter=Q(date_initiation__date=yesterday)),
        hier_frais=Sum('frais_transfert', filter=Q(date_initiation__date=yesterday)),
        
        # Montants mensuels
        ce_mois_fcfa=Sum('montant_fcfa', filter=Q(date_initiation__gte=current_month)),
        ce_mois_frais=Sum('frais_transfert', filter=Q(date_initiation__gte=current_month)),
        mois_precedent_fcfa=Sum('montant_fcfa', filter=Q(
            date_initiation__gte=last_month,
            date_initiation__lt=current_month
        )),
        mois_precedent_frais=Sum('frais_transfert', filter=Q(
            date_initiation__gte=last_month,
            date_initiation__lt=current_month
        ))
    )
    
    # === STATISTIQUES COMPLÈTES DES COLIS ET LOTS ===
    colis_stats = {
        # Colis par statut
        'total_colis': Colis.objects.count(),
        'en_attente': Colis.objects.filter(statut='en_attente').count(),
        'receptionnes_chine': Colis.objects.filter(statut='receptionne_chine').count(),
        'en_transit': Colis.objects.filter(statut='en_transit').count(),
        'arrives_mali': Colis.objects.filter(statut='arrive').count(),
        'livres': Colis.objects.filter(statut='livre').count(),
        'perdus': Colis.objects.filter(statut='perdu').count(),
        
        # Colis par mode de paiement
        'payes_chine': Colis.objects.filter(mode_paiement='paye_chine').count(),
        'payes_mali': Colis.objects.filter(mode_paiement='paye_mali').count(),
        'non_payes': Colis.objects.filter(mode_paiement='non_paye').count(),
        
        # Colis par type de transport
        'cargo': Colis.objects.filter(type_transport='cargo').count(),
        'express': Colis.objects.filter(type_transport='express').count(),
        'bateau': Colis.objects.filter(type_transport='bateau').count(),
    }
    
    # === STATISTIQUES DES LOTS ===
    lots_stats = {
        'total_lots': Lot.objects.count(),
        'ouverts': Lot.objects.filter(statut='ouvert').count(),
        'fermes': Lot.objects.filter(statut='ferme').count(),
        'expedies': Lot.objects.filter(statut='expedie').count(),
        'en_transit': Lot.objects.filter(statut='en_transit').count(),
        'arrives': Lot.objects.filter(statut='arrive').count(),
        'livres': Lot.objects.filter(statut='livre').count(),
    }
    
    # === VALEURS FINANCIÈRES DES COLIS ===
    valeurs_colis = Colis.objects.aggregate(
        # Valeur totale de tous les colis
        valeur_totale=Sum('prix_calcule'),
        
        # Valeur des colis en Chine (stock)
        valeur_stock_chine=Sum('prix_calcule', filter=Q(
            statut__in=['en_attente', 'receptionne_chine']
        )),
        
        # Valeur des colis en transit
        valeur_transit=Sum('prix_calcule', filter=Q(statut='en_transit')),
        
        # Valeur des colis arrivés au Mali
        valeur_arrives_mali=Sum('prix_calcule', filter=Q(statut='arrive')),
        
        # Valeur des colis livrés (chiffre d'affaires réalisé)
        valeur_livres=Sum('prix_calcule', filter=Q(statut='livre')),
        
        # Valeur des colis payés en Chine
        valeur_payes_chine=Sum('prix_calcule', filter=Q(mode_paiement='paye_chine')),
        
        # Valeur des colis à collecter au Mali
        valeur_a_collecter=Sum('prix_calcule', filter=Q(mode_paiement='paye_mali')),
    )
    
    # === STATISTIQUES DES PRIX DE TRANSPORT DES LOTS ===
    prix_lots = Lot.objects.aggregate(
        prix_transport_total=Sum('prix_transport'),
        prix_lots_en_transit=Sum('prix_transport', filter=Q(statut='en_transit')),
        prix_lots_expedies=Sum('prix_transport', filter=Q(statut='expedie')),
    )
    
    # === STATISTIQUES DES DÉPENSES ===
    try:
        depenses_stats = Depense.objects.aggregate(
            depenses_totales=Sum('montant'),
            depenses_ce_mois=Sum('montant', filter=Q(date_depense__gte=current_month)),
            depenses_hier=Sum('montant', filter=Q(date_depense=yesterday)),
            depenses_aujourd_hui=Sum('montant', filter=Q(date_depense=today)),
        )
        
        # Dépenses par type
        depenses_par_type = {}
        for type_dep, _ in Depense.TYPE_DEPENSE_CHOICES:
            depenses_par_type[type_dep] = Depense.objects.filter(
                type_depense=type_dep
            ).aggregate(total=Sum('montant'))['total'] or 0
    except:
        depenses_stats = {
            'depenses_totales': 0,
            'depenses_ce_mois': 0,
            'depenses_hier': 0,
            'depenses_aujourd_hui': 0,
        }
        depenses_par_type = {}
    
    # === STATISTIQUES DES AGENTS ===
    agents_stats = {
        'total_agents_chine': CustomUser.objects.filter(is_agent_chine=True).count(),
        'agents_chine_actifs': CustomUser.objects.filter(is_agent_chine=True, is_active=True).count(),
        'total_agents_mali': CustomUser.objects.filter(is_agent_mali=True).count(),
        'agents_mali_actifs': CustomUser.objects.filter(is_agent_mali=True, is_active=True).count(),
        'total_clients': CustomUser.objects.filter(is_client=True).count(),
        'clients_actifs': CustomUser.objects.filter(is_client=True, is_active=True).count(),
        'admins_chine': CustomUser.objects.filter(is_admin_chine=True).count(),
        'admins_mali': CustomUser.objects.filter(is_admin_mali=True).count(),
    }
    
    # === MÉTRIQUES DE PERFORMANCE ===
    # Calculs de ratios et indicateurs de performance
    performance_metrics = {
        'taux_livraison': round(
            (colis_stats['livres'] / colis_stats['total_colis'] * 100) 
            if colis_stats['total_colis'] > 0 else 0, 1
        ),
        'taux_paiement_chine': round(
            (colis_stats['payes_chine'] / colis_stats['total_colis'] * 100) 
            if colis_stats['total_colis'] > 0 else 0, 1
        ),
        'taux_reussite_transferts': round(
            (transferts_stats['confirmes'] / transferts_stats['total'] * 100) 
            if transferts_stats['total'] > 0 else 0, 1
        ),
        'ratio_lots_en_transit': round(
            (lots_stats['en_transit'] / lots_stats['total_lots'] * 100) 
            if lots_stats['total_lots'] > 0 else 0, 1
        ),
    }
    
    # Transferts récents (5 derniers)
    transferts_recents = TransfertArgent.objects.select_related(
        'admin_mali', 'admin_chine'
    ).order_by('-date_initiation')[:5]
    
    # Lots récents en transit (5 derniers)
    lots_recents = Lot.objects.filter(
        statut__in=['en_transit', 'expedie']
    ).order_by('-date_expedition')[:5]
    
    # Colis récents livrés (5 derniers)
    colis_recents = Colis.objects.filter(
        statut='livre'
    ).select_related('client__user').order_by('-date_modification')[:5]
    
    # Données pour graphiques (évolution des montants sur 6 mois)
    graphique_transferts = []
    graphique_colis = []
    
    for i in range(6):
        mois = today.replace(day=1) - timedelta(days=i*30)
        mois_suivant = mois.replace(day=28) + timedelta(days=4)
        mois_suivant = mois_suivant.replace(day=1)
        
        # Données transferts
        montant_transferts_mois = TransfertArgent.objects.filter(
            date_initiation__gte=mois,
            date_initiation__lt=mois_suivant
        ).aggregate(total=Sum('montant_fcfa'))['total'] or 0
        
        # Données colis
        valeur_colis_mois = Colis.objects.filter(
            date_creation__gte=mois,
            date_creation__lt=mois_suivant
        ).aggregate(total=Sum('prix_calcule'))['total'] or 0
        
        graphique_transferts.append({
            'mois': mois.strftime('%b %Y'),
            'montant': float(montant_transferts_mois)
        })
        
        graphique_colis.append({
            'mois': mois.strftime('%b %Y'),
            'valeur': float(valeur_colis_mois)
        })
    
    graphique_transferts.reverse()
    graphique_colis.reverse()
    
    context = {
        'title': 'Dashboard Admin Mali',
        'transferts_stats': transferts_stats,
        'montants_transferts': montants_transferts,
        'colis_stats': colis_stats,
        'lots_stats': lots_stats,
        'valeurs_colis': valeurs_colis,
        'prix_lots': prix_lots,
        'depenses_stats': depenses_stats,
        'depenses_par_type': depenses_par_type,
        'agents_stats': agents_stats,
        'performance_metrics': performance_metrics,
        'transferts_recents': transferts_recents,
        'lots_recents': lots_recents,
        'colis_recents': colis_recents,
        'graphique_transferts': graphique_transferts,
        'graphique_colis': graphique_colis,
        'today': today,
    }
    
    return render(request, 'admin_mali_app/dashboard.html', context)


@admin_mali_required
def transferts_list(request):
    """
    Liste des transferts d'argent avec filtres et pagination
    """
    from django.core.paginator import Paginator
    from datetime import datetime
    
    transferts = TransfertArgent.objects.select_related(
        'admin_mali', 'admin_chine'
    ).order_by('-date_initiation')
    
    # Filtres
    statut_filter = request.GET.get('statut')
    if statut_filter:
        transferts = transferts.filter(statut=statut_filter)
    
    # Filtre par date
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    if date_debut:
        try:
            date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
            transferts = transferts.filter(date_initiation__date__gte=date_debut_obj)
        except ValueError:
            pass
    
    if date_fin:
        try:
            date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
            transferts = transferts.filter(date_initiation__date__lte=date_fin_obj)
        except ValueError:
            pass
    
    # Pagination
    paginator = Paginator(transferts, 10)  # 10 transferts par page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'title': 'Transferts d\'Argent',
        'transferts': page_obj,
        'statut_filter': statut_filter,
        'status_choices': TransfertArgent.STATUS_CHOICES,
    }
    
    return render(request, 'admin_mali_app/transferts_list.html', context)


@admin_mali_required
def transfert_create(request):
    """
    Créer un nouveau transfert d'argent
    """
    if request.method == 'POST':
        try:
            # Récupération des données du formulaire
            montant_fcfa = request.POST.get('montant_fcfa')
            methode_transfert = request.POST.get('methode_transfert')
            destinataire_nom = request.POST.get('destinataire_nom')
            destinataire_telephone = request.POST.get('destinataire_telephone')
            motif_transfert = request.POST.get('motif_transfert', '')
            
            # Validation des données obligatoires
            if not all([montant_fcfa, methode_transfert, destinataire_nom, destinataire_telephone]):
                messages.error(request, "Tous les champs obligatoires doivent être remplis.")
                return render(request, 'admin_mali_app/transfert_form.html', {
                    'title': 'Créer un Transfert',
                    'form_data': request.POST
                })
            
            # Conversion et validation du montant
            try:
                montant_decimal = Decimal(montant_fcfa)
                if montant_decimal <= 0:
                    raise ValueError("Montant invalide")
            except (ValueError, TypeError):
                messages.error(request, "Le montant doit être un nombre positif.")
                return render(request, 'admin_mali_app/transfert_form.html', {
                    'title': 'Créer un Transfert',
                    'form_data': request.POST
                })
            
            # Calcul des frais de transfert (2% du montant)
            frais_transfert = montant_decimal * Decimal('0.02')
            
            # Génération du numéro de transfert unique
            import uuid
            numero_transfert = f"TF{timezone.now().strftime('%Y%m%d')}{str(uuid.uuid4())[:8].upper()}"
            
            # Création du transfert
            transfert = TransfertArgent.objects.create(
                numero_transfert=numero_transfert,
                montant_fcfa=montant_decimal,
                frais_transfert=frais_transfert,
                methode_transfert=methode_transfert,
                destinataire_nom=destinataire_nom,
                destinataire_telephone=destinataire_telephone,
                motif_transfert=motif_transfert,
                admin_mali=request.user,
                statut='initie'
            )
            
            messages.success(request, f"Transfert {numero_transfert} créé avec succès!")
            return redirect('admin_mali_app:transfert_detail', transfert_id=transfert.id)
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la création du transfert: {str(e)}")
            return render(request, 'admin_mali_app/transfert_form.html', {
                'title': 'Créer un Transfert',
                'form_data': request.POST
            })
    
    # Récupération des choix pour le formulaire
    choix_methodes = TransfertArgent.METHODE_CHOICES
    
    return render(request, 'admin_mali_app/transfert_form.html', {
        'title': 'Créer un Transfert',
        'choix_methodes': choix_methodes
    })


@admin_mali_required
def transfert_edit(request, transfert_id):
    """
    Modifier un transfert d'argent existant
    """
    transfert = get_object_or_404(TransfertArgent, id=transfert_id)
    
    # Vérifier si le transfert peut être modifié
    if transfert.statut in ['confirme_chine', 'annule']:
        messages.error(request, "Ce transfert ne peut plus être modifié.")
        return redirect('admin_mali_app:transfert_detail', transfert_id=transfert.id)
    
    if request.method == 'POST':
        try:
            # Récupération des données du formulaire
            montant_fcfa = request.POST.get('montant_fcfa')
            methode_transfert = request.POST.get('methode_transfert')
            destinataire_nom = request.POST.get('destinataire_nom')
            destinataire_telephone = request.POST.get('destinataire_telephone')
            motif_transfert = request.POST.get('motif_transfert', '')
            statut = request.POST.get('statut', transfert.statut)
            
            # Validation des données obligatoires
            if not all([montant_fcfa, methode_transfert, destinataire_nom, destinataire_telephone]):
                messages.error(request, "Tous les champs obligatoires doivent être remplis.")
                return render(request, 'admin_mali_app/transfert_form.html', {
                    'title': 'Modifier le Transfert',
                    'transfert': transfert,
                    'form_data': request.POST
                })
            
            # Conversion et validation du montant
            try:
                montant_decimal = Decimal(montant_fcfa)
                if montant_decimal <= 0:
                    raise ValueError("Montant invalide")
            except (ValueError, TypeError):
                messages.error(request, "Le montant doit être un nombre positif.")
                return render(request, 'admin_mali_app/transfert_form.html', {
                    'title': 'Modifier le Transfert',
                    'transfert': transfert,
                    'form_data': request.POST
                })
            
            # Calcul des frais de transfert (2% du montant)
            frais_transfert = montant_decimal * Decimal('0.02')
            
            # Mise à jour du transfert
            transfert.montant_fcfa = montant_decimal
            transfert.frais_transfert = frais_transfert
            transfert.methode_transfert = methode_transfert
            transfert.destinataire_nom = destinataire_nom
            transfert.destinataire_telephone = destinataire_telephone
            transfert.motif_transfert = motif_transfert
            transfert.statut = statut
            transfert.save()
            
            messages.success(request, f"Transfert {transfert.numero_transfert} modifié avec succès!")
            return redirect('admin_mali_app:transfert_detail', transfert_id=transfert.id)
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la modification du transfert: {str(e)}")
            return render(request, 'admin_mali_app/transfert_form.html', {
                'title': 'Modifier le Transfert',
                'transfert': transfert,
                'form_data': request.POST
            })
    
    # Récupération des choix pour le formulaire
    choix_methodes = TransfertArgent.METHODE_CHOICES
    choix_statuts = TransfertArgent.STATUS_CHOICES
    
    return render(request, 'admin_mali_app/transfert_form.html', {
        'title': 'Modifier le Transfert',
        'transfert': transfert,
        'choix_methodes': choix_methodes,
        'choix_statuts': choix_statuts,
        'is_edit': True
    })


@admin_mali_required
def transfert_detail(request, transfert_id):
    """
    Détail d'un transfert d'argent
    """
    transfert = get_object_or_404(TransfertArgent, id=transfert_id)
    
    context = {
        'title': f'Transfert {transfert.numero_transfert}',
        'transfert': transfert,
    }
    
    return render(request, 'admin_mali_app/transfert_detail.html', context)


@admin_mali_required
def agents_list(request):
    """
    Gestion des agents (Chine et Mali)
    """
    agents_chine = CustomUser.objects.filter(is_agent_chine=True).order_by('first_name')
    agents_mali = CustomUser.objects.filter(is_agent_mali=True).order_by('first_name')
    
    context = {
        'title': 'Gestion des Agents',
        'agents_chine': agents_chine,
        'agents_mali': agents_mali,
    }
    
    return render(request, 'admin_mali_app/agents_list.html', context)


@admin_mali_required
def tarifs_list(request):
    """
    Gestion des tarifs de transport
    """
    from reporting_app.models import ShippingPrice
    
    # Récupérer tous les tarifs
    tariffs = ShippingPrice.objects.all().order_by('-date_creation')
    
    # Filtres
    actif_filter = request.GET.get('actif')
    if actif_filter:
        tariffs = tariffs.filter(actif=actif_filter == 'true')
    
    pays_filter = request.GET.get('pays')
    if pays_filter:
        tariffs = tariffs.filter(pays_destination=pays_filter)
    
    context = {
        'title': 'Gestion des Tarifs',
        'tariffs': tariffs,
        'actif_filter': actif_filter,
        'pays_filter': pays_filter,
    }
    
    return render(request, 'admin_mali_app/tarifs_list.html', context)


@admin_mali_required
def rapports(request):
    """
    Rapports financiers et opérationnels avec filtres dynamiques
    """
    from django.db.models import Avg, Max, Min
    from datetime import datetime
    import calendar
    
    # Récupération des filtres
    type_rapport = request.GET.get('type_rapport', '')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    # Dates par défaut (30 derniers jours)
    today = timezone.now().date()
    if not date_debut:
        date_debut = today - timedelta(days=30)
    else:
        try:
            date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
        except ValueError:
            date_debut = today - timedelta(days=30)
    
    if not date_fin:
        date_fin = today
    else:
        try:
            date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError:
            date_fin = today
    
    rapport_data = None
    
    # Génération des données selon le type de rapport
    if type_rapport == 'transferts':
        rapport_data = generate_transferts_report(date_debut, date_fin)
    elif type_rapport == 'financier':
        rapport_data = generate_financial_report(date_debut, date_fin)
    elif type_rapport == 'agents':
        rapport_data = generate_agents_report(date_debut, date_fin)
    elif type_rapport == 'operationnel':
        rapport_data = generate_operational_report(date_debut, date_fin)
    
    context = {
        'title': 'Rapports',
        'rapport_data': rapport_data,
        'type_rapport': type_rapport,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    
    return render(request, 'admin_mali_app/rapports.html', context)


def generate_transferts_report(date_debut, date_fin):
    """
    Génère le rapport des transferts
    """
    transferts = TransfertArgent.objects.filter(
        date_initiation__date__gte=date_debut,
        date_initiation__date__lte=date_fin
    )
    
    total_transferts = transferts.count()
    total_montant = transferts.aggregate(total=Sum('montant_fcfa'))['total'] or 0
    montant_moyen = transferts.aggregate(avg=Avg('montant_fcfa'))['avg'] or 0
    transferts_en_cours = transferts.filter(statut='initie').count()
    
    return {
        'total_transferts': total_transferts,
        'total_montant': total_montant,
        'montant_moyen': montant_moyen,
        'transferts_en_cours': transferts_en_cours,
        'transferts': transferts.order_by('-date_initiation')[:50]  # Limiter à 50 pour l'affichage
    }


def generate_financial_report(date_debut, date_fin):
    """
    Génère le rapport financier
    """
    transferts = TransfertArgent.objects.filter(
        date_initiation__date__gte=date_debut,
        date_initiation__date__lte=date_fin
    )
    
    revenus_total = transferts.aggregate(total=Sum('montant_fcfa'))['total'] or 0
    commissions = transferts.aggregate(total=Sum('frais_transfert'))['total'] or 0
    benefice_net = commissions * Decimal('0.8')  # Approximation du bénéfice net
    
    # Données mensuelles pour le graphique
    mois_data = []
    revenus_mensuels = []
    
    # Générer les données des 6 derniers mois
    current_date = date_fin
    for i in range(6):
        mois_debut = current_date.replace(day=1)
        if mois_debut.month == 1:
            mois_fin = mois_debut.replace(year=mois_debut.year, month=12, day=31)
        else:
            next_month = mois_debut.replace(month=mois_debut.month - 1)
            mois_fin = (next_month.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        
        revenus_mois = TransfertArgent.objects.filter(
            date_initiation__date__gte=mois_debut,
            date_initiation__date__lte=mois_fin
        ).aggregate(total=Sum('montant_fcfa'))['total'] or 0
        
        mois_data.append(calendar.month_name[mois_debut.month])
        revenus_mensuels.append(float(revenus_mois))
        
        # Passer au mois précédent
        if current_date.month == 1:
            current_date = current_date.replace(year=current_date.year - 1, month=12)
        else:
            current_date = current_date.replace(month=current_date.month - 1)
    
    mois_data.reverse()
    revenus_mensuels.reverse()
    
    return {
        'revenus_total': revenus_total,
        'commissions': commissions,
        'benefice_net': benefice_net,
        'mois': mois_data,
        'revenus_mensuels': revenus_mensuels
    }


def generate_agents_report(date_debut, date_fin):
    """
    Génère le rapport des performances des agents
    """
    # Agents Mali avec leurs statistiques
    agents_mali = CustomUser.objects.filter(is_agent_mali=True, is_active=True)
    agents_data = []
    
    for agent in agents_mali:
        transferts_agent = TransfertArgent.objects.filter(
            admin_mali=agent,
            date_initiation__date__gte=date_debut,
            date_initiation__date__lte=date_fin
        )
        
        nb_transferts = transferts_agent.count()
        montant_total = transferts_agent.aggregate(total=Sum('montant_fcfa'))['total'] or 0
        commission = transferts_agent.aggregate(total=Sum('frais_transfert'))['total'] or 0
        
        # Calcul de performance basé sur le nombre de transferts et les montants
        performance = min(100, (nb_transferts * 10) + (float(montant_total) / 1000000 * 20))
        
        agents_data.append({
            'nom': agent.last_name,
            'prenom': agent.first_name,
            'nb_transferts': nb_transferts,
            'montant_total': montant_total,
            'commission': commission,
            'performance': round(performance, 1)
        })
    
    # Trier par performance décroissante
    agents_data.sort(key=lambda x: x['performance'], reverse=True)
    
    return {
        'agents': agents_data
    }


def generate_operational_report(date_debut, date_fin):
    """
    Génère le rapport opérationnel
    """
    transferts = TransfertArgent.objects.filter(
        date_initiation__date__gte=date_debut,
        date_initiation__date__lte=date_fin
    )
    
    total_transferts = transferts.count()
    transferts_reussis = transferts.filter(statut='confirme_chine').count()
    transferts_en_cours = transferts.filter(statut__in=['initie', 'envoye']).count()
    transferts_echecs = transferts.filter(statut='annule').count()
    
    # Calculs des métriques opérationnelles
    taux_reussite = round((transferts_reussis / total_transferts * 100) if total_transferts > 0 else 0, 1)
    temps_moyen = 45  # Temps moyen estimé en minutes
    pic_activite = "14h-16h"  # Pic d'activité estimé
    satisfaction = 85  # Taux de satisfaction estimé
    
    return {
        'taux_reussite': taux_reussite,
        'temps_moyen': temps_moyen,
        'pic_activite': pic_activite,
        'satisfaction': satisfaction,
        'transferts_reussis': transferts_reussis,
        'transferts_en_cours': transferts_en_cours,
        'transferts_echecs': transferts_echecs
    }


@admin_mali_required
def transfert_delete(request, transfert_id):
    """
    Supprimer un transfert d'argent
    """
    transfert = get_object_or_404(TransfertArgent, id=transfert_id)
    
    # Vérifier si le transfert peut être supprimé
    if transfert.statut in ['confirme_chine']:
        messages.error(request, "Un transfert confirmé ne peut pas être supprimé.")
        return redirect('admin_mali_app:transfert_detail', transfert_id=transfert.id)
    
    if request.method == 'POST':
        numero_transfert = transfert.numero_transfert
        transfert.delete()
        messages.success(request, f"Transfert {numero_transfert} supprimé avec succès!")
        return redirect('admin_mali_app:transferts_list')
    
    context = {
        'title': f'Supprimer le Transfert {transfert.numero_transfert}',
        'transfert': transfert,
    }
    
    return render(request, 'admin_mali_app/transfert_delete_confirm.html', context)


@admin_mali_required
def agent_create(request):
    """
    Créer un nouvel agent
    """
    if request.method == 'POST':
        try:
            # Récupération des données du formulaire
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            email = request.POST.get('email')
            phone_number = request.POST.get('phone_number')
            role = request.POST.get('role')  # 'agent_mali' ou 'agent_chine'
            
            # Validation des données obligatoires
            if not all([first_name, last_name, email, phone_number, role]):
                messages.error(request, "Tous les champs obligatoires doivent être remplis.")
                return render(request, 'admin_mali_app/agent_form.html', {
                    'title': 'Créer un Agent',
                    'form_data': request.POST
                })
            
            # Vérifier si l'email existe déjà
            if CustomUser.objects.filter(email=email).exists():
                messages.error(request, "Cet email est déjà utilisé.")
                return render(request, 'admin_mali_app/agent_form.html', {
                    'title': 'Créer un Agent',
                    'form_data': request.POST
                })
            
            # Création de l'utilisateur agent
            agent = CustomUser.objects.create_user(
                username=email,
                email=email,
                first_name=first_name,
                last_name=last_name,
                phone_number=phone_number,
                password='agent123',  # Mot de passe par défaut
            )
            
            # Assignation du rôle
            if role == 'agent_mali':
                agent.is_agent_mali = True
            elif role == 'agent_chine':
                agent.is_agent_chine = True
            
            agent.save()
            
            messages.success(request, f"Agent {first_name} {last_name} créé avec succès! Mot de passe par défaut: 'agent123'")
            return redirect('admin_mali_app:agents_list')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la création de l'agent: {str(e)}")
            return render(request, 'admin_mali_app/agent_form.html', {
                'title': 'Créer un Agent',
                'form_data': request.POST
            })
    
    return render(request, 'admin_mali_app/agent_form.html', {
        'title': 'Créer un Agent'
    })


@admin_mali_required
def agent_edit(request, agent_id):
    """
    Modifier un agent existant
    """
    agent = get_object_or_404(CustomUser, id=agent_id)
    
    if request.method == 'POST':
        try:
            # Récupération des données du formulaire
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            email = request.POST.get('email')
            phone_number = request.POST.get('phone_number')
            is_active = request.POST.get('is_active') == 'on'
            
            # Validation des données obligatoires
            if not all([first_name, last_name, email, phone_number]):
                messages.error(request, "Tous les champs obligatoires doivent être remplis.")
                return render(request, 'admin_mali_app/agent_form.html', {
                    'title': 'Modifier l\'Agent',
                    'agent': agent,
                    'form_data': request.POST
                })
            
            # Vérifier si l'email existe déjà (sauf pour cet agent)
            if CustomUser.objects.filter(email=email).exclude(id=agent.id).exists():
                messages.error(request, "Cet email est déjà utilisé.")
                return render(request, 'admin_mali_app/agent_form.html', {
                    'title': 'Modifier l\'Agent',
                    'agent': agent,
                    'form_data': request.POST
                })
            
            # Mise à jour de l'agent
            agent.first_name = first_name
            agent.last_name = last_name
            agent.email = email
            agent.username = email
            agent.phone_number = phone_number
            agent.is_active = is_active
            agent.save()
            
            messages.success(request, f"Agent {first_name} {last_name} modifié avec succès!")
            return redirect('admin_mali_app:agents_list')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la modification de l'agent: {str(e)}")
            return render(request, 'admin_mali_app/agent_form.html', {
                'title': 'Modifier l\'Agent',
                'agent': agent,
                'form_data': request.POST
            })
    
    return render(request, 'admin_mali_app/agent_form.html', {
        'title': 'Modifier l\'Agent',
        'agent': agent,
        'is_edit': True
    })


@admin_mali_required
def agent_delete(request, agent_id):
    """
    Supprimer un agent
    """
    agent = get_object_or_404(CustomUser, id=agent_id)
    
    if request.method == 'POST':
        nom_complet = f"{agent.first_name} {agent.last_name}"
        agent.delete()
        messages.success(request, f"Agent {nom_complet} supprimé avec succès!")
        return redirect('admin_mali_app:agents_list')
    
    context = {
        'title': f'Supprimer l\'Agent {agent.first_name} {agent.last_name}',
        'agent': agent,
    }
    
    return render(request, 'admin_mali_app/agent_delete_confirm.html', context)


@admin_mali_required
def tarif_create(request):
    """
    Créer un nouveau tarif de transport
    """
    from reporting_app.models import ShippingPrice
    
    if request.method == 'POST':
        try:
            # Récupération des données du formulaire
            nom_tarif = request.POST.get('nom_tarif')
            methode_calcul = request.POST.get('methode_calcul')
            prix_par_kilo = request.POST.get('prix_par_kilo')
            prix_par_m3 = request.POST.get('prix_par_m3')
            prix_forfaitaire = request.POST.get('prix_forfaitaire')
            poids_minimum = request.POST.get('poids_minimum', 0)
            poids_maximum = request.POST.get('poids_maximum')
            volume_minimum = request.POST.get('volume_minimum', 0)
            volume_maximum = request.POST.get('volume_maximum')
            pays_destination = request.POST.get('pays_destination')
            description = request.POST.get('description', '')
            actif = request.POST.get('actif') == 'on'
            date_debut = request.POST.get('date_debut')
            date_fin = request.POST.get('date_fin')
            
            # Validation des données obligatoires
            if not all([nom_tarif, methode_calcul, pays_destination, date_debut]):
                messages.error(request, "Tous les champs obligatoires doivent être remplis.")
                return redirect('admin_mali_app:tarifs')
            
            # Validation des prix selon la méthode
            if methode_calcul == 'par_kilo' and not prix_par_kilo:
                messages.error(request, "Le prix par kilo est obligatoire pour cette méthode.")
                return redirect('admin_mali_app:tarifs')
            
            if methode_calcul == 'par_metre_cube' and not prix_par_m3:
                messages.error(request, "Le prix par m3 est obligatoire pour cette méthode.")
                return redirect('admin_mali_app:tarifs')
            
            if methode_calcul == 'forfaitaire' and not prix_forfaitaire:
                messages.error(request, "Le prix forfaitaire est obligatoire pour cette méthode.")
                return redirect('admin_mali_app:tarifs')
            
            # Conversion des dates
            from datetime import datetime
            try:
                date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
                date_fin_obj = None
                if date_fin:
                    date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, "Format de date invalide.")
                return redirect('admin_mali_app:tarifs')
            
            # Création du tarif
            tarif = ShippingPrice.objects.create(
                nom_tarif=nom_tarif,
                methode_calcul=methode_calcul,
                prix_par_kilo=Decimal(prix_par_kilo) if prix_par_kilo else None,
                prix_par_m3=Decimal(prix_par_m3) if prix_par_m3 else None,
                prix_forfaitaire=Decimal(prix_forfaitaire) if prix_forfaitaire else None,
                poids_minimum=Decimal(poids_minimum) if poids_minimum else 0,
                poids_maximum=Decimal(poids_maximum) if poids_maximum else None,
                volume_minimum=Decimal(volume_minimum) if volume_minimum else 0,
                volume_maximum=Decimal(volume_maximum) if volume_maximum else None,
                pays_destination=pays_destination,
                description=description,
                actif=actif,
                date_debut=date_debut_obj,
                date_fin=date_fin_obj,
                cree_par=request.user
            )
            
            messages.success(request, f"Tarif '{nom_tarif}' créé avec succès!")
            return redirect('admin_mali_app:tarifs')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la création du tarif: {str(e)}")
            return redirect('admin_mali_app:tarifs')
    
    return redirect('admin_mali_app:tarifs')


@admin_mali_required
def export_rapport_excel(request):
    """
    Exporter les rapports en format Excel avec formatage professionnel
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from datetime import datetime
    
    # Récupération des filtres
    type_rapport = request.GET.get('type_rapport', 'financier')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    # Dates par défaut
    today = timezone.now().date()
    if not date_debut:
        date_debut = today - timedelta(days=30)
    else:
        try:
            date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
        except ValueError:
            date_debut = today - timedelta(days=30)
    
    if not date_fin:
        date_fin = today
    else:
        try:
            date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError:
            date_fin = today
    
    # Création du workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Styles pour le formatage
    header_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF6B35', end_color='FF6B35', fill_type='solid')
    
    subheader_font = Font(name='Arial', size=12, bold=True, color='2D3748')
    subheader_fill = PatternFill(start_color='F7FAFC', end_color='F7FAFC', fill_type='solid')
    
    data_font = Font(name='Arial', size=10)
    number_font = Font(name='Arial', size=10, bold=True)
    
    border = Border(
        left=Side(border_style='thin', color='E2E8F0'),
        right=Side(border_style='thin', color='E2E8F0'),
        top=Side(border_style='thin', color='E2E8F0'),
        bottom=Side(border_style='thin', color='E2E8F0')
    )
    
    if type_rapport == 'financier':
        ws.title = "Rapport Financier"
        rapport_data = generate_financial_report(date_debut, date_fin)
        
        # En-tête principal
        ws['A1'] = 'RAPPORT FINANCIER - TS AIR CARGO MALI'
        ws.merge_cells('A1:F1')
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Sous-titre avec période
        ws['A2'] = f'Période: {date_debut.strftime("%d/%m/%Y")} au {date_fin.strftime("%d/%m/%Y")}'
        ws.merge_cells('A2:F2')
        ws['A2'].font = subheader_font
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Date de génération
        ws['A3'] = f'Généré le: {datetime.now().strftime("%d/%m/%Y à %H:%M")}'
        ws.merge_cells('A3:F3')
        ws['A3'].font = data_font
        ws['A3'].alignment = Alignment(horizontal='center')
        
        # Résumé exécutif
        ws['A5'] = 'RÉSUMÉ EXÉCUTIF'
        ws['A5'].font = subheader_font
        ws['A5'].fill = subheader_fill
        
        # Données financières principales
        financial_data = [
            ['Indicateur', 'Valeur (FCFA)', 'Statut'],
            ['Revenus Total', rapport_data['revenus_total'], 'Positif'],
            ['Commissions Perçues', rapport_data['commissions'], 'Positif'],
            ['Bénéfice Net', rapport_data['benefice_net'], 'Positif' if rapport_data['benefice_net'] >= 0 else 'Négatif'],
        ]
        
        for row_idx, row_data in enumerate(financial_data, start=6):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if row_idx == 6:  # En-tête
                    cell.font = subheader_font
                    cell.fill = subheader_fill
                else:
                    if col_idx == 2:  # Colonnes numériques
                        cell.font = number_font
                        cell.number_format = '#,##0'
                    else:
                        cell.font = data_font
        
    elif type_rapport == 'transferts':
        ws.title = "Rapport Transferts"
        rapport_data = generate_transferts_report(date_debut, date_fin)
        
        # En-têtes
        ws['A1'] = 'RAPPORT TRANSFERTS'
        ws['A2'] = f'Période: {date_debut} au {date_fin}'
        ws['A4'] = 'Total Transferts'
        ws['B4'] = rapport_data['total_transferts']
        ws['A5'] = 'Montant Total (FCFA)'
        ws['B5'] = rapport_data['total_montant']
        ws['A6'] = 'Montant Moyen (FCFA)'
        ws['B6'] = rapport_data['montant_moyen']
        
        # Liste détaillée des transferts
        ws['A8'] = 'Numéro'
        ws['B8'] = 'Date'
        ws['C8'] = 'Montant'
        ws['D8'] = 'Destinataire'
        ws['E8'] = 'Statut'
        
        row = 9
        for transfert in rapport_data['transferts']:
            ws[f'A{row}'] = transfert.numero_transfert
            ws[f'B{row}'] = transfert.date_initiation.strftime('%d/%m/%Y')
            ws[f'C{row}'] = float(transfert.montant_fcfa)
            ws[f'D{row}'] = transfert.destinataire_nom
            ws[f'E{row}'] = transfert.get_statut_display()
            row += 1
    
    elif type_rapport == 'agents':
        ws.title = "Rapport Agents"
        rapport_data = generate_agents_report(date_debut, date_fin)
        
        # En-têtes
        ws['A1'] = 'RAPPORT PERFORMANCE AGENTS'
        ws['A2'] = f'Période: {date_debut} au {date_fin}'
        ws['A4'] = 'Nom'
        ws['B4'] = 'Prénom'
        ws['C4'] = 'Nb Transferts'
        ws['D4'] = 'Montant Total'
        ws['E4'] = 'Commission'
        ws['F4'] = 'Performance %'
        
        row = 5
        for agent in rapport_data['agents']:
            ws[f'A{row}'] = agent['nom']
            ws[f'B{row}'] = agent['prenom']
            ws[f'C{row}'] = agent['nb_transferts']
            ws[f'D{row}'] = float(agent['montant_total'])
            ws[f'E{row}'] = float(agent['commission'])
            ws[f'F{row}'] = agent['performance']
            row += 1
    
    # Préparation de la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="rapport_{type_rapport}_{date_debut}_{date_fin}.xlsx"'
    
    wb.save(response)
    return response


@admin_mali_required
def tarif_edit(request, tarif_id):
    """
    Modifier un tarif de transport existant
    """
    from reporting_app.models import ShippingPrice
    
    tarif = get_object_or_404(ShippingPrice, id=tarif_id)
    
    if request.method == 'POST':
        try:
            # Récupération des données du formulaire
            nom_tarif = request.POST.get('nom_tarif')
            methode_calcul = request.POST.get('methode_calcul')
            prix_par_kilo = request.POST.get('prix_par_kilo')
            prix_par_m3 = request.POST.get('prix_par_m3')
            prix_forfaitaire = request.POST.get('prix_forfaitaire')
            poids_minimum = request.POST.get('poids_minimum', 0)
            poids_maximum = request.POST.get('poids_maximum')
            volume_minimum = request.POST.get('volume_minimum', 0)
            volume_maximum = request.POST.get('volume_maximum')
            pays_destination = request.POST.get('pays_destination')
            description = request.POST.get('description', '')
            actif = request.POST.get('actif') == 'on'
            date_debut = request.POST.get('date_debut')
            date_fin = request.POST.get('date_fin')
            
            # Validation des données obligatoires
            if not all([nom_tarif, methode_calcul, pays_destination, date_debut]):
                messages.error(request, "Tous les champs obligatoires doivent être remplis.")
                return render(request, 'admin_mali_app/tarif_form.html', {
                    'title': 'Modifier le Tarif',
                    'tarif': tarif,
                    'form_data': request.POST
                })
            
            # Validation des prix selon la méthode
            if methode_calcul == 'par_kilo' and not prix_par_kilo:
                messages.error(request, "Le prix par kilo est obligatoire pour cette méthode.")
                return render(request, 'admin_mali_app/tarif_form.html', {
                    'title': 'Modifier le Tarif',
                    'tarif': tarif,
                    'form_data': request.POST
                })
            
            if methode_calcul == 'par_metre_cube' and not prix_par_m3:
                messages.error(request, "Le prix par m3 est obligatoire pour cette méthode.")
                return render(request, 'admin_mali_app/tarif_form.html', {
                    'title': 'Modifier le Tarif',
                    'tarif': tarif,
                    'form_data': request.POST
                })
            
            # Mise à jour du tarif
            tarif.nom_tarif = nom_tarif
            tarif.methode_calcul = methode_calcul
            tarif.prix_par_kilo = Decimal(prix_par_kilo) if prix_par_kilo else None
            tarif.prix_par_m3 = Decimal(prix_par_m3) if prix_par_m3 else None
            tarif.prix_forfaitaire = Decimal(prix_forfaitaire) if prix_forfaitaire else None
            tarif.poids_minimum = Decimal(poids_minimum) if poids_minimum else 0
            tarif.poids_maximum = Decimal(poids_maximum) if poids_maximum else None
            tarif.volume_minimum = Decimal(volume_minimum) if volume_minimum else 0
            tarif.volume_maximum = Decimal(volume_maximum) if volume_maximum else None
            tarif.pays_destination = pays_destination
            tarif.description = description
            tarif.actif = actif
            
            # Conversion des dates
            from datetime import datetime
            try:
                tarif.date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
                if date_fin:
                    tarif.date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
                else:
                    tarif.date_fin = None
            except ValueError:
                messages.error(request, "Format de date invalide.")
                return render(request, 'admin_mali_app/tarif_form.html', {
                    'title': 'Modifier le Tarif',
                    'tarif': tarif,
                    'form_data': request.POST
                })
            
            tarif.save()
            
            messages.success(request, f"Tarif '{nom_tarif}' modifié avec succès!")
            return redirect('admin_mali_app:tarifs')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la modification du tarif: {str(e)}")
            return render(request, 'admin_mali_app/tarif_form.html', {
                'title': 'Modifier le Tarif',
                'tarif': tarif,
                'form_data': request.POST
            })
    
    return render(request, 'admin_mali_app/tarif_form.html', {
        'title': 'Modifier le Tarif',
        'tarif': tarif,
        'is_edit': True
    })


@admin_mali_required
def tarif_detail(request, tarif_id):
    """
    Détail d'un tarif de transport
    """
    from reporting_app.models import ShippingPrice
    
    tarif = get_object_or_404(ShippingPrice, id=tarif_id)
    
    context = {
        'title': f'Tarif {tarif.nom_tarif}',
        'tarif': tarif,
    }
    
    return render(request, 'admin_mali_app/tarif_detail.html', context)


@admin_mali_required
def tarif_delete(request, tarif_id):
    """
    Supprimer un tarif de transport
    """
    from reporting_app.models import ShippingPrice
    
    tarif = get_object_or_404(ShippingPrice, id=tarif_id)
    
    if request.method == 'POST':
        nom_tarif = tarif.nom_tarif
        tarif.delete()
        messages.success(request, f"Tarif '{nom_tarif}' supprimé avec succès!")
        return redirect('admin_mali_app:tarifs')
    
    context = {
        'title': f'Supprimer le Tarif {tarif.nom_tarif}',
        'tarif': tarif,
    }
    
    return render(request, 'admin_mali_app/tarif_delete_confirm.html', context)


@admin_mali_required
def parametres(request):
    """
    Configuration et paramètres système
    """
    context = {
        'title': 'Paramètres Système',
    }
    
    return render(request, 'admin_mali_app/parametres.html', context)
